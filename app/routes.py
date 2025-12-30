from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, send_file, current_app
from flask_login import login_required, current_user
import re
from datetime import datetime
import os
import zipfile
import tempfile
import json
import secrets
from typing import Dict, Any, List
from werkzeug.utils import secure_filename

# Security: File upload validation
try:
    import magic
    MAGIC_AVAILABLE = True
except ImportError:
    MAGIC_AVAILABLE = False
    magic = None

try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False
    Image = None
try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

from .models import Match, MatchCategory, MatchResult, AppSettings, PhysicalMeasurement, Achievement, ClubHistory, TrainingCamp, PhysicalMetrics, Reference, SubscriptionStatus, Subscription, User
from .storage import StorageManager
from .utils import validate_match_data, parse_input_date, format_date_for_input
from .pdf import generate_season_pdf, generate_scout_pdf
from .phv_calculator import calculate_phv, validate_measurements_for_phv, calculate_predicted_adult_height, calculate_age_at_date
from .elite_benchmarks import get_elite_benchmarks_for_age, compare_to_elite
from .config import SUPPORT_EMAIL, SUBSCRIPTION_PRICING, CURRENT_YEAR

# Create blueprint
bp = Blueprint('main', __name__)

# Initialize storage
storage = StorageManager()

# Free tier limits
FREE_TIER_LIMITS = {
    'matches': 5,
    'achievements': 1,
    'physical_measurements': 2,
    'physical_metrics': 2,
    'references': 1
}

def is_admin_user():
    """Check if current user is admin"""
    admin_username = os.environ.get('ADMIN_USERNAME', '').strip()
    if not admin_username:
        return False
    try:
        from flask_login import current_user
        return current_user.is_authenticated and current_user.username == admin_username
    except:
        return False

def get_authenticated_user_id():
    """
    Get the authenticated user's ID from Flask-Login.
    SECURITY: Never trust client-provided user_id. Always use current_user.id
    """
    if not current_user.is_authenticated:
        raise ValueError("User is not authenticated")
    return current_user.id

def redirect_if_admin():
    """Redirect admin users to admin page if they try to access player data"""
    if is_admin_user():
        return redirect(url_for('main.admin_users'))
    return None

def check_subscription_and_limit(user_id, resource_type, current_count=None):
    """
    Check if user has active subscription and if they've reached free tier limits.
    Returns: (has_access, limit, current_count, error_message)
    """
    subscription = storage.get_subscription_by_user_id(user_id)
    has_active_subscription = subscription and subscription.status == SubscriptionStatus.ACTIVE
    
    if has_active_subscription:
        return (True, None, current_count, None)  # No limits for paid users
    
    # Free user - check limits
    limit = FREE_TIER_LIMITS.get(resource_type)
    if limit is None:
        return (True, None, current_count, None)  # No limit for this resource type
    
    # Get current count if not provided
    if current_count is None:
        if resource_type == 'matches':
            matches = storage.load_matches(user_id)
            current_count = len(matches)
        elif resource_type == 'achievements':
            achievements = storage.load_achievements(user_id)
            current_count = len(achievements)
        elif resource_type == 'physical_measurements':
            measurements = storage.load_physical_measurements(user_id)
            current_count = len(measurements)
        elif resource_type == 'physical_metrics':
            metrics = storage.load_physical_metrics(user_id)
            current_count = len(metrics)
        elif resource_type == 'references':
            references = storage.load_references(user_id)
            current_count = len(references)
    
    if current_count >= limit:
        error_msg = f"Free tier limit reached: {limit} {resource_type.replace('_', ' ')}. Subscribe to unlock unlimited entries."
        return (False, limit, current_count, error_msg)
    
    return (True, limit, current_count, None)


@bp.route('/test')
def test():
    """Test page to verify app is running"""
    return render_template('test.html')


@bp.route('/')
def homepage():
    """Homepage - Public landing page - data loaded client-side"""
    # If logged in as admin, redirect to admin page
    try:
        if current_user.is_authenticated:
            admin_username = os.environ.get('ADMIN_USERNAME', '').strip()
            current_username = current_user.username.strip() if current_user.username else ''
            if admin_username and current_username == admin_username:
                current_app.logger.info(f"Redirecting admin user '{current_username}' from homepage to admin page")
                return redirect(url_for('main.admin_users'))
    except:
        pass  # If not authenticated or error, continue to homepage
    
    # Return empty/default data - actual data will be loaded client-side
    default_settings = {
        'club_name': '',
        'player_name': '',
        'season_year': ''
    }
    return render_template('homepage.html', 
                         is_authenticated=False,  # Will be determined client-side
                         settings=default_settings,
                         matches=[],
                         achievements=[],
                         season_stats={'total_matches': 0, 'wins': 0, 'draws': 0, 'losses': 0, 'goals': 0, 'assists': 0, 'minutes': 0},
                         total_achievements=0)


@bp.route('/dashboard')
@login_required
def dashboard():
    """Main dashboard page - data loaded client-side"""
    # Redirect admin users to admin page immediately
    admin_username = os.environ.get('ADMIN_USERNAME', '').strip()
    current_username = current_user.username.strip() if current_user.username else ''
    if admin_username and current_username == admin_username:
        current_app.logger.info(f"Redirecting admin user '{current_username}' from dashboard to admin page")
        return redirect(url_for('main.admin_users'))
    
    # Return empty/default data - actual data will be loaded client-side
    default_settings = {
        'club_name': '',
        'player_name': '',
        'season_year': ''
    }
    return render_template('index.html', 
                         settings=default_settings,
                         matches=[],
                         achievements=[],
                         season_stats={'total_matches': 0, 'wins': 0, 'draws': 0, 'losses': 0, 'goals': 0, 'assists': 0, 'minutes': 0})


@bp.route('/matches')
@login_required
def matches():
    """Matches list page - data loaded client-side"""
    # Redirect admin users to admin page
    admin_username = os.environ.get('ADMIN_USERNAME', '').strip()
    if admin_username and current_user.username == admin_username:
        return redirect(url_for('main.admin_users'))
    
    # Return empty/default data - actual data will be loaded client-side
    default_settings = {
        'club_name': '',
        'player_name': '',
        'season_year': ''
    }
    return render_template('matches.html',
                     settings=default_settings,
                     completed_matches=[],
                     fixtures=[])


@bp.route('/matches', methods=['POST'])
@login_required
def create_match():
    """Create a new match"""
    user_id = current_user.id
    
    # Check subscription and limits
    has_access, limit, current_count, error_msg = check_subscription_and_limit(user_id, 'matches')
    if not has_access:
        return jsonify({'success': False, 'errors': [error_msg], 'limit_reached': True, 'current_count': current_count, 'limit': limit}), 403
    
    data = request.get_json()
    
    # Validate data
    errors = validate_match_data(data)
    if errors:
        return jsonify({'success': False, 'errors': errors}), 400
    
    try:
        # Parse date if it's in HTML format
        if 'date' in data and '-' in data['date']:
            data['date'] = parse_input_date(data['date'])
        
        # Handle clean sheets field
        clean_sheets = None
        if 'clean_sheets' in data and data['clean_sheets'] is not None and data['clean_sheets'] != '':
            try:
                clean_sheets = int(data['clean_sheets'])
            except (ValueError, TypeError):
                clean_sheets = None
        
        # Create match object
        match = Match(
            category=MatchCategory(data['category']),
            date=data['date'],
            opponent=data['opponent'],
            location=data['location'],
            result=MatchResult(data['result']) if data.get('result') else None,
            score=data.get('score', ''),
            brodie_goals=int(data.get('brodie_goals', 0)),
            brodie_assists=int(data.get('brodie_assists', 0)),
            clean_sheets=clean_sheets,
            minutes_played=int(data.get('minutes_played', 0)),
            notes=data.get('notes', ''),
            is_fixture=data.get('is_fixture', False),
            include_in_report=data.get('include_in_report', True)  # Default to True
        )
        
        # Save match
        match_id = storage.save_match(match, user_id)
        
        # Return the saved match object
        saved_match = storage.get_match(match_id, user_id)
        return jsonify({'success': True, 'match_id': match_id, 'match': saved_match.model_dump() if saved_match else None})
        
    except (ValueError, TypeError, KeyError) as e:
        return jsonify({'success': False, 'errors': [f'Invalid match data: {str(e)}']}), 400
    except Exception as e:
        return jsonify({'success': False, 'errors': [str(e)]}), 400


@bp.route('/matches/<match_id>', methods=['PUT'])
@login_required
def update_match(match_id):
    """Update an existing match"""
    user_id = current_user.id
    data = request.get_json()
    
    # Validate data
    errors = validate_match_data(data)
    if errors:
        return jsonify({'success': False, 'errors': errors}), 400
    
    try:
        # Get existing match
        existing_match = storage.get_match(match_id, user_id)
        if not existing_match:
            return jsonify({'success': False, 'errors': ['Match not found']}), 404
        
        # Parse date if it's in HTML format
        if 'date' in data and '-' in data['date']:
            data['date'] = parse_input_date(data['date'])
        
        # Handle clean sheets field
        clean_sheets = existing_match.clean_sheets
        if 'clean_sheets' in data:
            if data['clean_sheets'] is None or data['clean_sheets'] == '':
                clean_sheets = None
            else:
                try:
                    clean_sheets = int(data['clean_sheets'])
                except (ValueError, TypeError):
                    clean_sheets = None
        
        # Update match object
        updated_match = Match(
            id=match_id,
            category=MatchCategory(data['category']),
            date=data['date'],
            opponent=data['opponent'],
            location=data['location'],
            result=MatchResult(data['result']) if data.get('result') else None,
            score=data.get('score', ''),
            brodie_goals=int(data.get('brodie_goals', 0)),
            brodie_assists=int(data.get('brodie_assists', 0)),
            clean_sheets=clean_sheets,
            minutes_played=int(data.get('minutes_played', 0)),
            notes=data.get('notes', ''),
            is_fixture=data.get('is_fixture', False),
            include_in_report=data.get('include_in_report', getattr(existing_match, 'include_in_report', True))
        )
        
        # Save updated match
        storage.save_match(updated_match, user_id)
        
        # Return the updated match object
        saved_match = storage.get_match(match_id, user_id)
        return jsonify({'success': True, 'match': saved_match.model_dump() if saved_match else None})
        
    except (ValueError, TypeError, KeyError) as e:
        return jsonify({'success': False, 'errors': [f'Invalid match data: {str(e)}']}), 400
    except Exception as e:
        return jsonify({'success': False, 'errors': [str(e)]}), 400


@bp.route('/matches/<match_id>', methods=['DELETE'])
@login_required
def delete_match(match_id):
    """Delete a match"""
    user_id = current_user.id
    success = storage.delete_match(match_id, user_id)
    if success:
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'errors': ['Match not found']}), 404


@bp.route('/matches/<match_id>')
@login_required
def get_match(match_id):
    """Get a specific match"""
    user_id = current_user.id
    match = storage.get_match(match_id, user_id)
    if match:
        return jsonify({'success': True, 'match': match.model_dump()})
    else:
        return jsonify({'success': False, 'errors': ['Match not found']}), 404


@bp.route('/fixtures')
@login_required
def fixtures():
    """Get upcoming fixtures"""
    user_id = current_user.id
    fixtures = storage.get_fixtures(user_id)
    return jsonify({'success': True, 'fixtures': [f.model_dump() for f in fixtures]})


@bp.route('/pdf', methods=['POST'])
@login_required
def generate_pdf():
    """Generate PDF report - accepts data from client"""
    try:
        # Check subscription status - try multiple methods
        subscription = storage.get_subscription_by_user_id(current_user.id)
        
        # If no subscription found by user_id, check if there's exactly one active subscription
        # This handles cases where user_id format changed (e.g., client-side ID vs server-side ID)
        if not subscription:
            subscriptions = storage.load_subscriptions()
            active_subs = [s for s in subscriptions if s.get('status', '').lower() == 'active']
            
            # If there's exactly one active subscription, it's likely for the current user
            # Update it to use the correct user_id
            if len(active_subs) == 1:
                sub_data = active_subs[0]
                try:
                    # Create subscription object and update user_id
                    if 'status' in sub_data and isinstance(sub_data['status'], str):
                        sub_data['status'] = SubscriptionStatus(sub_data['status'].lower())
                    subscription = Subscription(**sub_data)
                    subscription.user_id = current_user.id
                    storage.save_subscription(subscription)
                    current_app.logger.info(f"Updated subscription user_id from {sub_data.get('user_id')} to {current_user.id}")
                except Exception as e:
                    current_app.logger.error(f"Error updating subscription: {e}")
                    subscription = None
        
        # Debug logging
        current_app.logger.info(f"PDF generation request for user {current_user.id}")
        current_app.logger.info(f"Subscription found: {subscription is not None}")
        if subscription:
            current_app.logger.info(f"Subscription status: {subscription.status}, type: {type(subscription.status)}")
            current_app.logger.info(f"Subscription status == ACTIVE: {subscription.status == SubscriptionStatus.ACTIVE}")
            current_app.logger.info(f"Subscription status value: {subscription.status.value if hasattr(subscription.status, 'value') else subscription.status}")
        
        # Check if subscription is active - simplified check
        is_active = False
        if subscription:
            # Get status value (works for both enum and string)
            status_value = subscription.status.value if hasattr(subscription.status, 'value') else str(subscription.status)
            is_active = status_value.lower() == SubscriptionStatus.ACTIVE.value.lower()
            current_app.logger.info(f"Status value: {status_value}, ACTIVE value: {SubscriptionStatus.ACTIVE.value}, is_active: {is_active}")
        
        if not subscription or not is_active:
            current_app.logger.warning(f"PDF generation blocked for user {current_user.id}: subscription={subscription is not None}, is_active={is_active}")
            return jsonify({
                'success': False,
                'errors': ['PDF generation is a premium feature. Please subscribe to unlock this feature.'],
                'debug': {
                    'has_subscription': subscription is not None,
                    'status': subscription.status if subscription else None,
                    'user_id': current_user.id
                }
            }), 403
        
        data = request.get_json() if request.is_json else {}
        user_id = current_user.id
        
        # Get data from request (client-side storage)
        matches_data = data.get('matches', [])
        settings_data = data.get('settings', {})
        physical_measurements_data = data.get('physical_measurements', [])
        physical_metrics_data = data.get('physical_metrics', [])
        period = data.get('period', 'all_time')
        
        # Validate period
        valid_periods = ['all_time', 'season', '12_months', '6_months', '3_months', 'last_month']
        if period not in valid_periods:
            period = 'all_time'
        
        # Load settings from server (most up-to-date, includes player profile data)
        # Merge with client settings to ensure we have all data
        try:
            server_settings = storage.load_settings(user_id)
            if server_settings:
                # Merge server settings (source of truth) with client settings
                # Server settings take priority, but client settings can fill in gaps
                server_dict = server_settings.model_dump()
                # Only update with client settings that are not None/empty in server settings
                # AND the client value is not None/empty
                for key, value in settings_data.items():
                    if (key not in server_dict or 
                        server_dict[key] is None or 
                        server_dict[key] == '' or
                        (isinstance(server_dict[key], list) and len(server_dict[key]) == 0)):
                        # Only use client value if it's not None/empty
                        if value is not None and value != '' and not (isinstance(value, list) and len(value) == 0):
                            server_dict[key] = value
                settings_data = server_dict
        except Exception as e:
            # If loading from server fails, use client settings
            print(f"Warning: Could not load settings from server: {e}")
            pass
        
        # Convert data to model objects
        from .models import Match, AppSettings, PhysicalMeasurement, PhysicalMetrics
        
        matches = [Match(**m) for m in matches_data]
        settings = AppSettings(**settings_data) if settings_data else AppSettings()
        
        # Convert physical measurement dates from YYYY-MM-DD to dd MMM yyyy format if needed
        cleaned_measurements_data = []
        for pm in physical_measurements_data:
            cleaned_pm = pm.copy()
            if 'date' in cleaned_pm and cleaned_pm['date']:
                date_str = str(cleaned_pm['date']).strip()
                # Check if date is in YYYY-MM-DD format (starts with 4 digits, has dashes)
                # Try to parse as YYYY-MM-DD first
                if re.match(r'^\d{4}-\d{2}-\d{2}', date_str):
                    try:
                        cleaned_pm['date'] = parse_input_date(date_str)
                        print(f"Converted date: {date_str} -> {cleaned_pm['date']}")
                    except Exception as e:
                        print(f"Warning: Could not convert date {date_str}: {e}")
                        continue  # Skip this measurement if date conversion fails
                # If already in correct format, validate it
                elif not re.match(r'^\d{1,2} [A-Za-z]{3} \d{4}', date_str):
                    # Not in expected format, try to convert anyway
                    try:
                        cleaned_pm['date'] = parse_input_date(date_str)
                        print(f"Converted date (fallback): {date_str} -> {cleaned_pm['date']}")
                    except Exception as e:
                        print(f"Warning: Could not convert date {date_str}: {e}")
                        continue
            cleaned_measurements_data.append(cleaned_pm)
        
        physical_measurements = [PhysicalMeasurement(**pm) for pm in cleaned_measurements_data]
        
        # Convert physical metrics dates from YYYY-MM-DD to dd MMM yyyy format if needed
        # Also clean numeric fields (convert empty strings to None)
        cleaned_metrics_data = []
        numeric_fields = [
            'sprint_speed_ms', 'sprint_speed_kmh', 'sprint_10m_sec', 'sprint_20m_sec', 'sprint_30m_sec',
            'vertical_jump_cm', 'standing_long_jump_cm', 'countermovement_jump_cm',
            'agility_time_sec', 'yo_yo_test_level', 'beep_test_level',
            'bench_press_kg', 'squat_kg', 'deadlift_kg',
            'vo2_max', 'sit_and_reach_cm'
        ]
        integer_fields = ['max_heart_rate', 'resting_heart_rate']
        
        for pm in physical_metrics_data:
            if not pm or not isinstance(pm, dict):
                continue
            cleaned_pm = {}
            # Initialize ALL numeric fields to None first
            for field in numeric_fields:
                cleaned_pm[field] = None
            for field in integer_fields:
                cleaned_pm[field] = None
            
            for key, value in pm.items():
                # Handle numeric fields - convert empty strings to None
                if key in numeric_fields:
                    # Check for empty string, None, or whitespace-only string
                    if value is None or value == '' or (isinstance(value, str) and value.strip() == ''):
                        cleaned_pm[key] = None
                    else:
                        try:
                            # Try to convert to float
                            cleaned_pm[key] = float(value)
                        except (ValueError, TypeError):
                            cleaned_pm[key] = None
                # Handle integer fields - convert empty strings to None
                elif key in integer_fields:
                    # Check for empty string, None, or whitespace-only string
                    if value is None or value == '' or (isinstance(value, str) and value.strip() == ''):
                        cleaned_pm[key] = None
                    else:
                        try:
                            # Try to convert to int
                            cleaned_pm[key] = int(value)
                        except (ValueError, TypeError):
                            cleaned_pm[key] = None
                # Handle date field
                elif key == 'date' and value:
                    date_str = str(value).strip()
                    if date_str:
                        # Check if date is already in correct format (dd MMM yyyy)
                        if not re.match(r'^\d{1,2} [A-Za-z]{3} \d{4}', date_str):
                            # Check if it's YYYY-MM-DD format
                            if re.match(r'^\d{4}-\d{2}-\d{2}', date_str):
                                try:
                                    converted_date = parse_input_date(date_str)
                                    if converted_date != date_str and re.match(r'^\d{1,2} [A-Za-z]{3} \d{4}', converted_date):
                                        cleaned_pm[key] = converted_date
                                    else:
                                        print(f"Warning: Date conversion failed for {date_str}")
                                        continue  # Skip this metric
                                except Exception as e:
                                    print(f"Error: Could not convert date {date_str}: {e}")
                                    continue  # Skip this metric
                            else:
                                # Try to convert anyway
                                try:
                                    converted_date = parse_input_date(date_str)
                                    if converted_date != date_str and re.match(r'^\d{1,2} [A-Za-z]{3} \d{4}', converted_date):
                                        cleaned_pm[key] = converted_date
                                    else:
                                        print(f"Warning: Could not convert date format: {date_str}")
                                        continue
                                except Exception as e:
                                    print(f"Error: Could not convert date {date_str}: {e}")
                                    continue
                        else:
                            cleaned_pm[key] = date_str
                    else:
                        continue  # Skip if date is empty
                else:
                    # Copy other fields as-is (id, notes, etc.) - but skip numeric fields
                    if key not in numeric_fields + integer_fields:
                        cleaned_pm[key] = value
            
            # CRITICAL: Ensure all numeric fields are explicitly set to None (never empty strings)
            for field in numeric_fields + integer_fields:
                # Always set to None if missing, empty string, or invalid
                if field not in cleaned_pm:
                    cleaned_pm[field] = None
                elif cleaned_pm[field] == '':
                    cleaned_pm[field] = None
                elif isinstance(cleaned_pm[field], str):
                    if not cleaned_pm[field].strip():
                        cleaned_pm[field] = None
                    else:
                        # Try to convert string to number
                        try:
                            if field in integer_fields:
                                cleaned_pm[field] = int(cleaned_pm[field])
                            else:
                                cleaned_pm[field] = float(cleaned_pm[field])
                        except (ValueError, TypeError):
                            cleaned_pm[field] = None
            
            # Only add if we have a valid date
            if 'date' in cleaned_pm and cleaned_pm['date']:
                cleaned_metrics_data.append(cleaned_pm)
        
        print(f"Processed {len(cleaned_metrics_data)} physical metrics out of {len(physical_metrics_data)}")
        
        # AGGRESSIVE final cleanup: create completely new dicts with only valid values
        final_cleaned_metrics = []
        for pm in cleaned_metrics_data:
            final_pm = {}
            # Copy required fields
            if 'id' in pm:
                final_pm['id'] = pm['id']
            if 'date' in pm and pm['date']:
                final_pm['date'] = pm['date']
            if 'notes' in pm:
                final_pm['notes'] = pm.get('notes')
            
            # Process ALL numeric fields - explicitly convert empty strings to None
            for field in numeric_fields:
                value = pm.get(field)
                if value is None:
                    final_pm[field] = None
                elif value == '':
                    final_pm[field] = None
                elif isinstance(value, str):
                    if value.strip() == '':
                        final_pm[field] = None
                    else:
                        try:
                            final_pm[field] = float(value)
                        except (ValueError, TypeError):
                            final_pm[field] = None
                elif isinstance(value, (int, float)):
                    final_pm[field] = float(value)
                else:
                    final_pm[field] = None
            
            for field in integer_fields:
                value = pm.get(field)
                if value is None:
                    final_pm[field] = None
                elif value == '':
                    final_pm[field] = None
                elif isinstance(value, str):
                    if value.strip() == '':
                        final_pm[field] = None
                    else:
                        try:
                            final_pm[field] = int(value)
                        except (ValueError, TypeError):
                            final_pm[field] = None
                elif isinstance(value, int):
                    final_pm[field] = value
                elif isinstance(value, float):
                    final_pm[field] = int(value)
                else:
                    final_pm[field] = None
            
            final_cleaned_metrics.append(final_pm)
        
        # Debug: print first metric to verify
        if final_cleaned_metrics:
            print(f"Final cleaned metric sample - keys: {list(final_cleaned_metrics[0].keys())}")
            for field in ['vertical_jump_cm', 'countermovement_jump_cm', 'agility_time_sec', 'yo_yo_test_level', 'bench_press_kg']:
                if field in final_cleaned_metrics[0]:
                    val = final_cleaned_metrics[0][field]
                    print(f"  {field}: {repr(val)} (type: {type(val).__name__})")
        
        # Create PhysicalMetrics - ONLY include fields with valid values (skip empty strings entirely)
        physical_metrics = []
        for i, pm in enumerate(final_cleaned_metrics):
            try:
                # Build dict with ONLY valid values - don't include empty strings at all
                safe_pm = {}
                
                # Required fields
                if 'id' in pm:
                    safe_pm['id'] = pm['id']
                if 'date' in pm and pm['date']:
                    safe_pm['date'] = pm['date']
                if 'notes' in pm and pm.get('notes'):
                    safe_pm['notes'] = pm['notes']
                
                # Only include numeric fields if they have valid (non-empty) values
                for field in numeric_fields:
                    val = pm.get(field)
                    # Skip if None, empty string, or whitespace-only string
                    if val is None or val == '' or (isinstance(val, str) and not val.strip()):
                        continue  # Don't include this field - Pydantic will use default None
                    # Try to convert to float
                    try:
                        safe_pm[field] = float(val)
                    except (ValueError, TypeError):
                        continue  # Skip invalid values
                
                for field in integer_fields:
                    val = pm.get(field)
                    # Skip if None, empty string, or whitespace-only string
                    if val is None or val == '' or (isinstance(val, str) and not val.strip()):
                        continue  # Don't include this field - Pydantic will use default None
                    # Try to convert to int
                    try:
                        safe_pm[field] = int(val)
                    except (ValueError, TypeError):
                        continue  # Skip invalid values
                
                # Create PhysicalMetrics - empty string fields won't be in the dict, so Pydantic uses None
                physical_metrics.append(PhysicalMetrics(**safe_pm))
            except Exception as e:
                print(f"Error creating PhysicalMetrics for metric {i}: {e}")
                print(f"  Problematic data: {pm}")
                import traceback
                traceback.print_exc()
                # Skip this metric
                continue
        
        # Create output directory
        output_dir = os.path.join(current_app.root_path, '..', 'output')
        os.makedirs(output_dir, exist_ok=True)
        
        # Generate PDF with period filter
        pdf_path = generate_season_pdf(matches, settings, output_dir, physical_measurements, physical_metrics, period=period)
        
        # Return the PDF file directly
        return send_file(
            pdf_path,
            as_attachment=True,
            download_name=os.path.basename(pdf_path),
            mimetype='application/pdf'
        )
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error generating PDF: {error_details}")
        current_app.logger.error(f"PDF generation error: {error_details}")
        return jsonify({'success': False, 'errors': [str(e)]}), 500


@bp.route('/export')
@login_required
def export_data():
    """Export all data as Excel file"""
    temp_file = None
    try:
        user_id = current_user.id
        # Get all data for this user
        data = storage.export_data(user_id)
        
        # Create temporary Excel file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Helper function to add a sheet with data
        def add_sheet(sheet_name, data_list, headers):
            if not data_list:
                return
            ws = wb.create_sheet(title=sheet_name)
            
            # Add headers with styling
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            # Add data rows
            for row_idx, item in enumerate(data_list, 2):
                if isinstance(item, dict):
                    for col_idx, header in enumerate(headers, 1):
                        value = item.get(header, '')
                        # Convert None to empty string
                        if value is None:
                            value = ''
                        ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Add Matches sheet
        if data.get('matches'):
            matches_headers = ['id', 'date', 'opponent', 'venue', 'result', 'goals', 'assists', 
                             'minutes_played', 'category', 'season', 'notes', 'clean_sheets']
            matches_data = []
            for match in data['matches']:
                if isinstance(match, dict):
                    matches_data.append(match)
                else:
                    matches_data.append(match.model_dump() if hasattr(match, 'model_dump') else match.dict())
            add_sheet('Matches', matches_data, matches_headers)
        
        # Add Settings sheet
        if data.get('settings'):
            settings_dict = data['settings']
            if hasattr(settings_dict, 'model_dump'):
                settings_dict = settings_dict.model_dump()
            elif hasattr(settings_dict, 'dict'):
                settings_dict = settings_dict.dict()
            
            ws = wb.create_sheet(title='Settings')
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            ws.cell(row=1, column=1, value='Setting').fill = header_fill
            ws.cell(row=1, column=1).font = header_font
            ws.cell(row=1, column=2, value='Value').fill = header_fill
            ws.cell(row=1, column=2).font = header_font
            
            row = 2
            for key, value in settings_dict.items():
                if value is not None:
                    ws.cell(row=row, column=1, value=str(key))
                    # Handle list values
                    if isinstance(value, list):
                        ws.cell(row=row, column=2, value=', '.join(str(v) for v in value))
                    else:
                        ws.cell(row=row, column=2, value=str(value))
                    row += 1
        
        # Add Physical Measurements sheet
        if data.get('physical_measurements'):
            measurements_headers = ['id', 'date', 'height_cm', 'weight_kg', 'notes', 'include_in_report']
            measurements_data = []
            for m in data['physical_measurements']:
                if isinstance(m, dict):
                    measurements_data.append(m)
                else:
                    measurements_data.append(m.model_dump() if hasattr(m, 'model_dump') else m.dict())
            add_sheet('Physical Measurements', measurements_data, measurements_headers)
        
        # Add Physical Metrics sheet
        if data.get('physical_metrics'):
            metrics_headers = ['id', 'date', 'sprint_speed_ms', 'sprint_speed_kmh', 'sprint_10m_sec', 
                             'sprint_20m_sec', 'sprint_30m_sec', 'vertical_jump_cm', 'standing_long_jump_cm',
                             'countermovement_jump_cm', 'agility_time_sec', 'beep_test_level', 'vo2_max',
                             'include_in_report']
            metrics_data = []
            for m in data['physical_metrics']:
                if isinstance(m, dict):
                    metrics_data.append(m)
                else:
                    metrics_data.append(m.model_dump() if hasattr(m, 'model_dump') else m.dict())
            add_sheet('Physical Metrics', metrics_data, metrics_headers)
        
        # Add Achievements sheet
        if data.get('achievements'):
            achievements_headers = ['id', 'date', 'title', 'category', 'season', 'goals', 'assists', 
                                   'minutes_played', 'clean_sheets', 'notes']
            achievements_data = []
            for a in data['achievements']:
                if isinstance(a, dict):
                    achievements_data.append(a)
                else:
                    achievements_data.append(a.model_dump() if hasattr(a, 'model_dump') else a.dict())
            add_sheet('Achievements', achievements_data, achievements_headers)
        
        # Add Club History sheet
        if data.get('club_history'):
            club_headers = ['id', 'club_name', 'season', 'age_group', 'position', 'achievements']
            club_data = []
            for c in data['club_history']:
                if isinstance(c, dict):
                    club_data.append(c)
                else:
                    club_data.append(c.model_dump() if hasattr(c, 'model_dump') else c.dict())
            add_sheet('Club History', club_data, club_headers)
        
        # Add Training Camps sheet
        if data.get('training_camps'):
            camp_headers = ['id', 'camp_name', 'organizer', 'location', 'start_date', 'end_date', 
                          'age_group', 'focus_area']
            camp_data = []
            for c in data['training_camps']:
                if isinstance(c, dict):
                    camp_data.append(c)
                else:
                    camp_data.append(c.model_dump() if hasattr(c, 'model_dump') else c.dict())
            add_sheet('Training Camps', camp_data, camp_headers)
        
        # Add References sheet
        if data.get('references'):
            ref_headers = ['id', 'name', 'position', 'organization', 'email', 'phone', 'relationship', 'notes']
            ref_data = []
            for r in data['references']:
                if isinstance(r, dict):
                    ref_data.append(r)
                else:
                    ref_data.append(r.model_dump() if hasattr(r, 'model_dump') else r.dict())
            add_sheet('References', ref_data, ref_headers)
        
        # Save workbook
        wb.save(temp_file.name)
        
        return send_file(temp_file.name, 
                        as_attachment=True, 
                        download_name='FutureElite_Export.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    except (IOError, OSError) as e:
        if temp_file and os.path.exists(temp_file.name):
            os.unlink(temp_file.name)
        return jsonify({'success': False, 'errors': [f'Error creating export: {str(e)}']}), 500
    except Exception as e:
        if temp_file and os.path.exists(temp_file.name):
            os.unlink(temp_file.name)
        return jsonify({'success': False, 'errors': [str(e)]}), 500


@bp.route('/import', methods=['POST'])
@login_required
def import_data():
    """Import data from Excel or ZIP file"""
    temp_file = None
    try:
        user_id = current_user.id
        if 'file' not in request.files:
            return jsonify({'success': False, 'errors': ['No file provided']}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'errors': ['No file selected']}), 400
        
        # Validate filename to prevent path traversal
        filename = os.path.basename(file.filename)
        is_excel = filename.endswith('.xlsx') or filename.endswith('.xls')
        is_zip = filename.endswith('.zip')
        
        if not (is_excel or is_zip):
            return jsonify({'success': False, 'errors': ['File must be an Excel file (.xlsx) or ZIP file (.zip)']}), 400
        
        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
        file.save(temp_file.name)
        
        # Security: Check file size (max 10MB)
        file_size = os.path.getsize(temp_file.name)
        if file_size > 10 * 1024 * 1024:
            os.unlink(temp_file.name)
            return jsonify({'success': False, 'errors': ['File too large. Maximum size is 10MB']}), 400
        
        # Import from Excel file
        if is_excel:
            if not EXCEL_SUPPORT:
                os.unlink(temp_file.name)
                return jsonify({'success': False, 'errors': ['Excel support not available. Please install openpyxl: pip install openpyxl']}), 400
            
            workbook = openpyxl.load_workbook(temp_file.name, data_only=True)
            
            # Initialize data structures
            matches_data = []
            settings_data = {}
            physical_measurements_data = []
            physical_metrics_data = []
            achievements_data = []
            club_history_data = []
            training_camps_data = []
            references_data = []
            
            # Helper function to read sheet data
            def read_sheet(sheet_name, headers=None):
                if sheet_name not in workbook.sheetnames:
                    return []
                ws = workbook[sheet_name]
                data = []
                if headers:
                    # Find header row
                    header_row = None
                    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                        if row and row[0] and str(row[0]).strip() in headers:
                            header_row = row_idx
                            break
                    
                    if header_row:
                        # Get headers from that row
                        header_map = {}
                        for col_idx, header in enumerate(ws[header_row], 1):
                            if header:
                                header_map[col_idx] = str(header).strip()
                        
                        # Read data rows
                        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                            if not any(row):  # Skip empty rows
                                continue
                            item = {}
                            for col_idx, value in enumerate(row, 1):
                                if col_idx in header_map and value is not None:
                                    item[header_map[col_idx]] = value
                            if item:
                                data.append(item)
                else:
                    # Read as key-value pairs (for Settings sheet)
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0] and row[1] is not None:
                            key = str(row[0]).strip()
                            value = row[1]
                            # Convert string representations of lists back to lists
                            if isinstance(value, str) and value.startswith('[') and value.endswith(']'):
                                try:
                                    import ast
                                    value = ast.literal_eval(value)
                                except:
                                    pass
                            settings_data[key] = value
                return data
            
            # Read each sheet
            matches_data = read_sheet('Matches', ['id', 'date', 'opponent', 'venue', 'result', 'goals', 'assists', 'minutes_played', 'category', 'season', 'notes', 'clean_sheets'])
            read_sheet('Settings')  # Settings is key-value format
            physical_measurements_data = read_sheet('Physical Measurements', ['id', 'date', 'height_cm', 'weight_kg', 'notes', 'include_in_report'])
            physical_metrics_data = read_sheet('Physical Metrics', ['id', 'date', 'sprint_speed_ms', 'sprint_speed_kmh', 'sprint_10m_sec', 'sprint_20m_sec', 'sprint_30m_sec', 'vertical_jump_cm', 'standing_long_jump_cm', 'countermovement_jump_cm', 'agility_time_sec', 'beep_test_level', 'vo2_max', 'include_in_report'])
            achievements_data = read_sheet('Achievements', ['id', 'date', 'title', 'category', 'season', 'goals', 'assists', 'minutes_played', 'clean_sheets', 'notes'])
            club_history_data = read_sheet('Club History', ['id', 'club_name', 'season', 'age_group', 'position', 'achievements'])
            training_camps_data = read_sheet('Training Camps', ['id', 'camp_name', 'organizer', 'location', 'start_date', 'end_date', 'age_group', 'focus_area'])
            references_data = read_sheet('References', ['id', 'name', 'position', 'organization', 'email', 'phone', 'relationship', 'notes'])
            
            # Prepare import data
            import_data = {
                'matches': matches_data,
                'settings': settings_data,
                'physical_measurements': physical_measurements_data,
                'club_history': club_history_data,
                'training_camps': training_camps_data,
                'physical_metrics': physical_metrics_data,
                'achievements': achievements_data,
                'references': references_data
            }
            
            success = storage.import_data(import_data, user_id)
            
            # Clean up
            if temp_file and os.path.exists(temp_file.name):
                os.unlink(temp_file.name)
            
            if success:
                return jsonify({'success': True, 'message': 'Data imported successfully from Excel file'})
            else:
                return jsonify({'success': False, 'errors': ['Failed to import data']}), 400
        
        # Import from ZIP file (existing functionality)
        # Security: ZIP bomb protection - check uncompressed size and file count
        MAX_UNCOMPRESSED_SIZE = 50 * 1024 * 1024  # 50MB
        MAX_FILES_IN_ZIP = 100
        MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB per file
        
        # Extract and import data
        with zipfile.ZipFile(temp_file.name, 'r') as zip_file:
            # Security: Check number of files
            file_list = zip_file.namelist()
            if len(file_list) > MAX_FILES_IN_ZIP:
                os.unlink(temp_file.name)
                return jsonify({'success': False, 'errors': [f'Too many files in archive. Maximum is {MAX_FILES_IN_ZIP} files.']}), 400
            
            # Security: Check uncompressed size
            total_size = 0
            for file_info in zip_file.infolist():
                total_size += file_info.file_size
                if total_size > MAX_UNCOMPRESSED_SIZE:
                    os.unlink(temp_file.name)
                    return jsonify({'success': False, 'errors': [f'Archive uncompressed size too large. Maximum is {MAX_UNCOMPRESSED_SIZE // (1024*1024)}MB.']}), 400
                if file_info.file_size > MAX_FILE_SIZE:
                    os.unlink(temp_file.name)
                    return jsonify({'success': False, 'errors': [f'File {file_info.filename} is too large. Maximum is {MAX_FILE_SIZE // (1024*1024)}MB per file.']}), 400
            
            # Security: Check compression ratio (prevent ZIP bombs)
            if file_size > 0:
                compression_ratio = total_size / file_size
                if compression_ratio > 100:  # Suspicious compression ratio
                    os.unlink(temp_file.name)
                    return jsonify({'success': False, 'errors': ['Suspicious compression ratio detected. File may be a ZIP bomb.']}), 400
            
            # Read files with size limits
            def read_zip_file_safe(zip_file, filename, max_size=MAX_FILE_SIZE):
                """Safely read a file from ZIP with size limit"""
                try:
                    file_data = zip_file.read(filename)
                    if len(file_data) > max_size:
                        raise ValueError(f"File {filename} exceeds size limit")
                    return file_data.decode('utf-8')
                except KeyError:
                    return None
            
            # Read matches data
            matches_json = read_zip_file_safe(zip_file, 'matches.json')
            if matches_json is None:
                os.unlink(temp_file.name)
                return jsonify({'success': False, 'errors': ['Required file matches.json not found in archive']}), 400
            matches_data = json.loads(matches_json)
            
            # Read settings data
            settings_json = read_zip_file_safe(zip_file, 'settings.json')
            if settings_json is None:
                os.unlink(temp_file.name)
                return jsonify({'success': False, 'errors': ['Required file settings.json not found in archive']}), 400
            settings_data = json.loads(settings_json)
            
            # Read physical measurements data (if exists)
            physical_measurements_data = []
            try:
                pm_json = read_zip_file_safe(zip_file, 'physical_measurements.json')
                if pm_json:
                    physical_measurements_data = json.loads(pm_json)
            except (KeyError, ValueError):
                pass
            
            # Read training camps data (if exists)
            training_camps_data = []
            try:
                tc_json = read_zip_file_safe(zip_file, 'training_camps.json')
                if tc_json:
                    training_camps_data = json.loads(tc_json)
            except (KeyError, ValueError):
                pass
            
            # Read physical metrics data (if exists)
            physical_metrics_data = []
            try:
                pm_json = read_zip_file_safe(zip_file, 'physical_metrics.json')
                if pm_json:
                    physical_metrics_data = json.loads(pm_json)
            except (KeyError, ValueError):
                pass
            
            # Read club history data (if exists)
            club_history_data = []
            try:
                ch_json = read_zip_file_safe(zip_file, 'club_history.json')
                if ch_json:
                    club_history_data = json.loads(ch_json)
            except (KeyError, ValueError):
                pass
        
        # Import data - assign user_id to all imported data
        import_data = {
            'matches': matches_data,
            'settings': settings_data,
            'physical_measurements': physical_measurements_data,
            'club_history': club_history_data,
            'training_camps': training_camps_data,
            'physical_metrics': physical_metrics_data
        }
        
        success = storage.import_data(import_data, user_id)
        
        # Clean up
        if temp_file and os.path.exists(temp_file.name):
            os.unlink(temp_file.name)
        
        if success:
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'errors': ['Failed to import data']}), 400
        
    except (zipfile.BadZipFile, json.JSONDecodeError, Exception) as e:
        if temp_file and os.path.exists(temp_file.name):
            os.unlink(temp_file.name)
        error_msg = str(e)
        if 'openpyxl' in error_msg.lower() or 'excel' in error_msg.lower():
            return jsonify({'success': False, 'errors': [f'Invalid Excel file format: {error_msg}']}), 400
        return jsonify({'success': False, 'errors': [f'Invalid file format: {error_msg}']}), 400
    except Exception as e:
        if temp_file and os.path.exists(temp_file.name):
            os.unlink(temp_file.name)
        return jsonify({'success': False, 'errors': [str(e)]}), 500


@bp.route('/download-excel-template')
def download_excel_template():
    """Download Excel template for match import"""
    if not EXCEL_SUPPORT:
        return jsonify({'success': False, 'errors': ['Excel support not available. Please install openpyxl: pip install openpyxl']}), 400
    
    try:
        # Create a new workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Matches"
        
        # Set column headers (exact format expected)
        headers = [
            'Opponent',
            'Location',
            'Date',
            'Category',
            'Score',
            'Goals',
            'Assists',
            'Minutes',
            'Notes'
        ]
        sheet.append(headers)
        
        # Style the header row
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = alignment
        except ImportError:
            # Styles not available, skip styling
            pass
        
        # Add example rows with different formats
        examples = [
            ['Altrajz', 'Al-Farabi', '28/08/2025', 'Pre-Season Friendly', '1 - 1', 0, 0, 15, ''],
            ['Faith Academy', 'Al-Farabi', '04/09/2025', 'Pre-Season Friendly', '2 - 2', 0, 0, 20, ''],
            ['Ole Academy', 'Al-Farabi', '11/09/2025', 'Pre-Season Friendly', '3 - 2', 1, 0, 20, ''],
            ['Aspire', 'Al-Farabi', '18/09/2025', 'Pre-Season Friendly', '1 - 1', 1, 0, 20, ''],
            ['Stars', 'Al-Farabi', '25/09/2025', 'Pre-Season Friendly', '2 - 2', 0, 0, 20, ''],
        ]
        
        for example in examples:
            sheet.append(example)
        
        # Add instructions row
        sheet.append([])
        sheet.append(['INSTRUCTIONS:', '', '', '', '', '', '', '', ''])
        instructions = [
            ['', 'Required columns: Opponent, Date', '', '', '', '', '', '', ''],
            ['', 'Date format: DD/MM/YYYY (e.g., 28/08/2025) or DD-MM-YYYY', '', '', '', '', '', '', ''],
            ['', 'Category: Pre-Season Friendly or League', '', '', '', '', '', '', ''],
            ['', 'Score format: "7 - 5" (our score - their score)', '', '', '', '', '', '', ''],
            ['', 'Goals, Assists, Minutes: Numbers only', '', '', '', '', '', '', ''],
            ['', 'Delete example rows before importing your data', '', '', '', '', '', '', ''],
        ]
        for instruction in instructions:
            sheet.append(instruction)
        
        # Auto-adjust column widths
        column_widths = {
            'A': 20,  # Opponent
            'B': 15,  # Location
            'C': 12,  # Date
            'D': 20,  # Category
            'E': 12,  # Score
            'F': 8,   # Goals
            'G': 8,   # Assists
            'H': 10,  # Minutes
            'I': 30,  # Notes
        }
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width
        
        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        workbook.save(temp_file.name)
        temp_file.close()
        
        # Return the file
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name='Match_Import_Template.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'success': False, 'errors': [f'Error creating template: {str(e)}']}), 500


@bp.route('/import-excel', methods=['POST'])
@login_required
def import_excel():
    """Import matches from Excel file or full backup - returns matches to client for saving"""
    if not EXCEL_SUPPORT:
        return jsonify({'success': False, 'errors': ['Excel support not available. Please install openpyxl: pip install openpyxl']}), 400
    
    temp_file = None
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'errors': ['No file provided']}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'errors': ['No file selected']}), 400
        
        # Validate filename to prevent path traversal
        filename = os.path.basename(file.filename)
        # Check file extension
        if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
            return jsonify({'success': False, 'errors': ['File must be an Excel file (.xlsx or .xls)']}), 400
        
        # Save file temporarily first
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        file.save(temp_file.name)
        
        # Check if this is a full backup (has multiple sheets) or match-only import
        workbook_check = openpyxl.load_workbook(temp_file.name, data_only=True)
        is_full_backup = len(workbook_check.sheetnames) > 1 or 'Settings' in workbook_check.sheetnames or 'Physical Measurements' in workbook_check.sheetnames
        
        # If it's a full backup, use the main import route which handles full backups
        if is_full_backup:
            # Close the workbook check
            workbook_check.close()
            # The temp_file will be handled by import_data()
            # We need to recreate the file object for import_data
            from werkzeug.datastructures import FileStorage
            import shutil
            temp_file_path = temp_file.name
            temp_file.close()
            
            # Create a new file object for import_data
            with open(temp_file_path, 'rb') as f:
                file_storage = FileStorage(
                    stream=f,
                    filename=filename,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                # Temporarily replace request.files['file'] with our file
                original_file = request.files.get('file')
                request.files['file'] = file_storage
                try:
                    result = import_data()
                    return result
                finally:
                    # Restore original file
                    if original_file:
                        request.files['file'] = original_file
                    # Clean up temp file
                    if os.path.exists(temp_file_path):
                        os.unlink(temp_file_path)
        
        # Get import mode
        import_mode = request.form.get('import_mode', 'replace')  # 'replace' or 'append'
        
        # Save file temporarily
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        file.save(temp_file.name)
        
        # Check file size (max 10MB for Excel import)
        file_size = os.path.getsize(temp_file.name)
        if file_size > 10 * 1024 * 1024:
            os.unlink(temp_file.name)
            return jsonify({'success': False, 'errors': ['File too large. Maximum size is 10MB']}), 400
        
        # Security: Parse Excel file in data-only mode to prevent formula injection
        workbook = openpyxl.load_workbook(temp_file.name, data_only=True)
        sheet = workbook.active
        
        # Security: Helper function to sanitize cell values and prevent formula injection
        def sanitize_cell_value(cell_value):
            """Sanitize cell value to prevent formula injection"""
            if cell_value is None:
                return None
            # Convert to string and strip
            value = str(cell_value).strip()
            # Security: Reject formulas (cells starting with =, +, -, @)
            if value and value[0] in ['=', '+', '-', '@']:
                # Log potential formula injection attempt
                current_app.logger.warning(f"Potential formula injection detected: {value[:50]}")
                return None  # Reject formulas
            return value
        
        # Find header row (look for "Opponent" or similar)
        header_row = None
        for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row and any(cell and 'opponent' in str(cell).lower().strip() for cell in row if cell):
                header_row = idx
                break
        
        # If not found, try first row as header
        if not header_row:
            header_row = 1
        
        # Read headers - handle None values and normalize, sanitize to prevent formula injection
        headers = []
        for cell in sheet[header_row]:
            if cell is None or cell.value is None:
                headers.append('')
            else:
                # Security: Sanitize header to prevent formula injection
                header_str = sanitize_cell_value(cell.value)
                headers.append(header_str if header_str else '')
        
        # Map column names to our fields (flexible matching)
        col_map = {}
        found_headers = []
        for idx, header in enumerate(headers):
            if not header:
                continue
            
            # Normalize header: lowercase, strip whitespace, normalize whitespace
            header_clean = str(header).lower().strip()
            header_clean = ' '.join(header_clean.split())  # Normalize multiple spaces to single space
            found_headers.append(f"'{header}'")
            
            # Opponent column - try multiple variations
            if 'opponent' in header_clean and 'opponent' not in col_map:
                col_map['opponent'] = idx
            elif header_clean in ['opp', 'vs', 'versus', 'against'] and 'opponent' not in col_map:
                col_map['opponent'] = idx
            
            # Location column
            if 'location' in header_clean and 'location' not in col_map:
                col_map['location'] = idx
            elif header_clean in ['venue', 'place', 'stadium', 'ground'] and 'location' not in col_map:
                col_map['location'] = idx
            
            # Date column - try multiple variations
            if 'date' in header_clean and 'date' not in col_map:
                col_map['date'] = idx
            elif header_clean in ['match date', 'game date', 'played', 'when'] and 'date' not in col_map:
                col_map['date'] = idx
            
            # Category column
            if ('category' in header_clean or 'type' in header_clean) and 'category' not in col_map:
                col_map['category'] = idx
            elif header_clean in ['match type', 'competition'] and 'category' not in col_map:
                col_map['category'] = idx
            
            # Score column
            if 'score' in header_clean and 'score' not in col_map:
                col_map['score'] = idx
            elif header_clean in ['result', 'final score', 'ft'] and 'score' not in col_map:
                col_map['score'] = idx
            
            # Goals column (check for "brodie goals" or just "goals" or "g")
            if 'goal' in header_clean and 'goals' not in col_map:
                if 'brodie' in header_clean or header_clean in ['g', 'goals', 'goal', 'brodie goals', 'brodie goal']:
                    col_map['goals'] = idx
            
            # Assists column
            if 'assist' in header_clean and 'assists' not in col_map:
                if 'brodie' in header_clean or header_clean in ['a', 'assists', 'assist', 'brodie assists', 'brodie assist']:
                    col_map['assists'] = idx
            
            # Minutes column
            if ('minute' in header_clean or 'min' in header_clean) and 'minutes' not in col_map:
                if 'brodie' in header_clean or header_clean in ['min', 'minutes', 'mins', 'playing minutes', 'time', 'played']:
                    col_map['minutes'] = idx
            
            # Notes column
            if 'note' in header_clean and 'notes' not in col_map:
                col_map['notes'] = idx
            elif header_clean in ['comment', 'comments', 'remarks'] and 'notes' not in col_map:
                col_map['notes'] = idx
        
        # Required columns - provide detailed error message
        if 'opponent' not in col_map or 'date' not in col_map:
            missing = []
            if 'opponent' not in col_map:
                missing.append('Opponent')
            if 'date' not in col_map:
                missing.append('Date')
            
            error_msg = f'Excel file must contain "{", ".join(missing)}" column(s). '
            error_msg += f'Found headers: {", ".join(found_headers[:10])}'  # Show first 10 headers
            if temp_file and os.path.exists(temp_file.name):
                os.unlink(temp_file.name)
            return jsonify({'success': False, 'errors': [error_msg]}), 400
        
        # Parse rows
        imported_matches = []
        errors = []
        match_counter = 0  # Counter to ensure unique IDs
        rows_checked = 0
        last_row_idx = header_row
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
            rows_checked += 1
            last_row_idx = row_idx
            # Skip completely empty rows
            if not row or not any(cell is not None and str(cell).strip() for cell in row if cell is not None):
                continue
            
            try:
                # Extract data - handle None and empty values better, sanitize to prevent formula injection
                opponent_idx = col_map.get('opponent')
                if opponent_idx is not None and opponent_idx < len(row):
                    opponent_raw = row[opponent_idx]
                    opponent = sanitize_cell_value(opponent_raw) if opponent_raw is not None else ''
                else:
                    opponent = ''
                
                # Skip if opponent is empty or is the header text or instruction rows
                if not opponent or opponent.lower() in ['opponent', 'opp', 'vs', 'versus', 'against', '']:
                    continue
                
                # Skip instruction rows (from template)
                if opponent.upper().startswith('INSTRUCTIONS') or opponent.upper().startswith('REQUIRED'):
                    continue
                
                # Extract other fields with better None handling, sanitize to prevent formula injection
                location_idx = col_map.get('location')
                if location_idx is not None and location_idx < len(row) and row[location_idx] is not None:
                    location = sanitize_cell_value(row[location_idx]) or 'Unknown'
                else:
                    location = 'Unknown'
                
                date_idx = col_map.get('date')
                if date_idx is not None and date_idx < len(row):
                    date_cell = sanitize_cell_value(row[date_idx])
                else:
                    date_cell = None
                
                category_idx = col_map.get('category')
                if category_idx is not None and category_idx < len(row) and row[category_idx] is not None:
                    category = sanitize_cell_value(row[category_idx]) or 'Pre-Season Friendly'
                else:
                    category = 'Pre-Season Friendly'
                
                score_idx = col_map.get('score')
                if score_idx is not None and score_idx < len(row) and row[score_idx] is not None:
                    score = sanitize_cell_value(row[score_idx])
                else:
                    score = None
                
                goals_idx = col_map.get('goals')
                if goals_idx is not None and goals_idx < len(row):
                    goals = row[goals_idx]
                else:
                    goals = 0
                
                assists_idx = col_map.get('assists')
                if assists_idx is not None and assists_idx < len(row):
                    assists = row[assists_idx]
                else:
                    assists = 0
                
                minutes_idx = col_map.get('minutes')
                if minutes_idx is not None and minutes_idx < len(row):
                    minutes = row[minutes_idx]
                else:
                    minutes = 0
                
                notes_idx = col_map.get('notes')
                if notes_idx is not None and notes_idx < len(row) and row[notes_idx] is not None:
                    notes = sanitize_cell_value(row[notes_idx]) or ''
                else:
                    notes = ''
                
                # Parse date - try multiple formats including Excel serial numbers
                date_str = None
                if date_cell is None:
                    errors.append(f"Row {row_idx}: Missing date (opponent: {opponent})")
                    continue
                elif isinstance(date_cell, datetime):
                    date_str = date_cell.strftime("%d %b %Y")
                elif isinstance(date_cell, (int, float)):
                    # Excel date serial number
                    try:
                        from openpyxl.utils import datetime_from_excel
                        date_obj = datetime_from_excel(date_cell)
                        date_str = date_obj.strftime("%d %b %Y")
                    except:
                        # Try as days since 1900
                        try:
                            from datetime import timedelta
                            base_date = datetime(1899, 12, 30)
                            date_obj = base_date + timedelta(days=int(date_cell))
                            date_str = date_obj.strftime("%d %b %Y")
                        except Exception as e:
                            errors.append(f"Row {row_idx}: Invalid date number '{date_cell}' (opponent: {opponent}): {str(e)}")
                            continue
                elif isinstance(date_cell, str):
                    date_cell_clean = date_cell.strip()
                    # Try multiple date formats
                    date_formats = [
                        "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y",
                        "%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d",
                        "%d %b %Y", "%d %B %Y",
                        "%b %d, %Y", "%B %d, %Y",
                        "%d/%m/%y", "%d-%m-%y"
                    ]
                    date_parsed = False
                    for fmt in date_formats:
                        try:
                            date_obj = datetime.strptime(date_cell_clean, fmt)
                            date_str = date_obj.strftime("%d %b %Y")
                            date_parsed = True
                            break
                        except ValueError:
                            continue
                    
                    if not date_parsed:
                        errors.append(f"Row {row_idx}: Invalid date format '{date_cell}' (opponent: {opponent})")
                        continue
                else:
                    errors.append(f"Row {row_idx}: Invalid date type '{type(date_cell)}' (opponent: {opponent})")
                    continue
                
                # Parse numeric values
                try:
                    goals = int(float(goals)) if goals else 0
                except (ValueError, TypeError):
                    goals = 0
                
                try:
                    assists = int(float(assists)) if assists else 0
                except (ValueError, TypeError):
                    assists = 0
                
                try:
                    minutes = int(float(minutes)) if minutes else 0
                except (ValueError, TypeError):
                    minutes = 0
                
                # Format score
                if score and score.strip() and score.lower() not in ['none', 'n/a', '']:
                    # Handle different score formats
                    score = score.replace(' ', '')
                    if '-' in score:
                        parts = score.split('-')
                        if len(parts) == 2:
                            score = f"{parts[0].strip()} - {parts[1].strip()}"
                    formatted_score = score
                else:
                    formatted_score = None
                
                # Determine result from score
                result = None
                if formatted_score:
                    try:
                        parts = formatted_score.split(' - ')
                        if len(parts) == 2:
                            our_score = int(parts[0].strip())
                            their_score = int(parts[1].strip())
                            if our_score > their_score:
                                result = MatchResult.WIN
                            elif our_score < their_score:
                                result = MatchResult.LOSS
                            else:
                                result = MatchResult.DRAW
                    except (ValueError, IndexError):
                        pass
                
                # Parse category
                try:
                    if category:
                        category_lower = category.lower()
                        if 'league' in category_lower:
                            match_category = MatchCategory.LEAGUE
                        elif 'friendly' in category_lower and 'pre-season' not in category_lower:
                            match_category = MatchCategory.FRIENDLY
                        elif 'pre-season' in category_lower or 'pre season' in category_lower:
                            match_category = MatchCategory.PRE_SEASON_FRIENDLY
                        else:
                            match_category = MatchCategory.PRE_SEASON_FRIENDLY
                    else:
                        match_category = MatchCategory.PRE_SEASON_FRIENDLY
                except (AttributeError, TypeError):
                    match_category = MatchCategory.PRE_SEASON_FRIENDLY
                
                # Create match with unique ID (use counter to ensure uniqueness)
                match_counter += 1
                base_id = datetime.now().strftime("%Y%m%d_%H%M%S")
                unique_id = f"{base_id}_{match_counter:04d}"
                
                match = Match(
                    id=unique_id,
                    category=match_category,
                    date=date_str,
                    opponent=opponent,
                    location=location,
                    result=result,
                    score=formatted_score,
                    brodie_goals=goals,
                    brodie_assists=assists,
                    minutes_played=minutes,
                    notes=notes,
                    is_fixture=False
                )
                
                imported_matches.append(match)
                
            except Exception as e:
                # Try to get opponent name for better error message
                try:
                    opponent_for_error = opponent if 'opponent' in locals() else 'unknown'
                    errors.append(f"Row {row_idx} (opponent: {opponent_for_error}): {str(e)}")
                except:
                    errors.append(f"Row {row_idx}: {str(e)}")
                continue
        
        # Clean up temp file
        if temp_file and os.path.exists(temp_file.name):
            os.unlink(temp_file.name)
        
        if not imported_matches:
            error_msg = 'No valid matches found in Excel file.'
            if errors:
                error_msg += f' Found {len(errors)} error(s).'
            return jsonify({'success': False, 'errors': [error_msg] + errors[:20]}), 400
        
        # Convert matches to dictionaries for JSON response
        matches_data = [m.model_dump() for m in imported_matches]
        
        response = {
            'success': True,
            'imported': len(imported_matches),
            'matches': matches_data,  # Return matches to client
            'import_mode': import_mode,  # Return mode so client knows what to do
            'rows_checked': rows_checked,
            'errors': errors[:20] if errors else []  # Show up to 20 errors
        }
        
        if errors:
            response['warning'] = f'Imported {len(imported_matches)} matches, but {len(errors)} row(s) had errors.'
        elif len(imported_matches) == 0:
            response['warning'] = 'No matches were imported. Please check that your Excel file has data rows with valid Opponent and Date values.'
        
        return jsonify(response)
        
    except (IOError, OSError, ValueError) as e:
        if temp_file and os.path.exists(temp_file.name):
            os.unlink(temp_file.name)
        return jsonify({'success': False, 'errors': [f'Invalid Excel file: {str(e)}']}), 400
    except Exception as e:
        if temp_file and os.path.exists(temp_file.name):
            os.unlink(temp_file.name)
        return jsonify({'success': False, 'errors': [f'Error importing Excel: {str(e)}']}), 500


@bp.route('/settings')
@login_required
def settings_page():
    """Settings page - data loaded client-side"""
    # Redirect admin users to admin page
    admin_username = os.environ.get('ADMIN_USERNAME', '').strip()
    if admin_username and current_user.username == admin_username:
        return redirect(url_for('main.admin_users'))
    
    # Return empty/default data - actual data will be loaded client-side
    default_settings = {
        'club_name': '',
        'player_name': '',
        'season_year': '',
        'league_name': '',
        'contact_email': '',
        'date_of_birth': None,
        'height_cm': None,
        'weight_kg': None,
        'phv_date': None,
        'phv_age': None,
        'position': '',
        'dominant_foot': '',
        'highlight_reel_urls': []
    }
    return render_template('settings.html', settings=default_settings)


@bp.route('/api/settings')
@login_required
def get_settings():
    """Get current settings (API endpoint)"""
    # Redirect admin users - return empty settings
    admin_username = os.environ.get('ADMIN_USERNAME', '').strip()
    if admin_username and current_user.username == admin_username:
        return jsonify({'success': True, 'settings': {}})
    
    user_id = current_user.id
    settings = storage.load_settings(user_id)
    return jsonify({'success': True, 'settings': settings.model_dump()})


@bp.route('/settings', methods=['POST'])
@login_required
def update_settings():
    """Update settings"""
    try:
        user_id = current_user.id
        data = request.get_json()
        
        # Load existing settings to preserve fields that might not be in the form
        existing_settings = storage.load_settings(user_id)
        existing_dict = existing_settings.model_dump()
        
        # Start with existing settings, but we'll handle special fields separately
        merged_data = {**existing_dict}
        
        # Parse date fields from incoming data if they're in HTML format (YYYY-MM-DD)
        # This must happen BEFORE merging to avoid validation errors
        date_fields = ['date_of_birth', 'phv_date']
        for field in date_fields:
            if field in data and data[field] is not None and data[field] != '':
                date_value = str(data[field]).strip()
                # Check if it's in HTML format (YYYY-MM-DD)
                if '-' in date_value and len(date_value) == 10:
                    try:
                        # Validate it's a valid date
                        datetime.strptime(date_value, "%Y-%m-%d")
                        # Convert to our format
                        data[field] = parse_input_date(date_value)
                    except ValueError:
                        # Invalid date format, set to None
                        data[field] = None
                # If it's already in our format (dd MMM yyyy), validate and keep it
                elif len(date_value) > 10:
                    try:
                        datetime.strptime(date_value, "%d %b %Y")
                        # Already in correct format, keep as is
                        pass
                    except ValueError:
                        # Invalid format, set to None
                        data[field] = None
            elif field in data and (data[field] == '' or data[field] is None):
                # Empty string or None, set to None
                data[field] = None
        
        # Handle special fields that need explicit handling
        # First, merge all regular fields from data (dates are now converted)
        # Treat social_media_links like any other field - simple merge
        for key, value in data.items():
            if key not in ['highlight_reel_urls', 'player_photo_path']:
                # For all regular fields including social_media_links and contact_email
                # Just merge directly - same logic as contact_email, position, etc.
                merged_data[key] = value
        
        # Handle highlight_reel_urls - always update if provided
        if 'highlight_reel_urls' in data:
            if isinstance(data['highlight_reel_urls'], str):
                try:
                    parsed = json.loads(data['highlight_reel_urls']) if data['highlight_reel_urls'] else []
                    merged_data['highlight_reel_urls'] = parsed if isinstance(parsed, list) else []
                except:
                    merged_data['highlight_reel_urls'] = existing_dict.get('highlight_reel_urls', [])
            elif isinstance(data['highlight_reel_urls'], list):
                merged_data['highlight_reel_urls'] = data['highlight_reel_urls']
            else:
                merged_data['highlight_reel_urls'] = existing_dict.get('highlight_reel_urls', [])
        else:
            merged_data['highlight_reel_urls'] = existing_dict.get('highlight_reel_urls', [])
        
        # Ensure social_media_links is always a dict (same as how other fields are handled)
        if 'social_media_links' not in merged_data or merged_data['social_media_links'] is None:
            merged_data['social_media_links'] = {}
        elif not isinstance(merged_data['social_media_links'], dict):
            # If it's not a dict, try to convert or default to empty dict
            merged_data['social_media_links'] = {}
        
        # Always preserve player_photo_path - it's managed separately via upload endpoint
        # Only update if explicitly provided and non-empty
        if 'player_photo_path' in data and data.get('player_photo_path'):
            merged_data['player_photo_path'] = data['player_photo_path']
        else:
            merged_data['player_photo_path'] = existing_dict.get('player_photo_path')
        
        # Create settings object
        settings = AppSettings(**merged_data)
        
        # Save settings
        storage.save_settings(settings, user_id)
        
        # If date_of_birth was provided and we have measurements, try to calculate PHV automatically
        # Only for paid users
        subscription = storage.get_subscription_by_user_id(user_id)
        has_active_subscription = subscription and subscription.status == SubscriptionStatus.ACTIVE
        
        if settings.date_of_birth and has_active_subscription:
            try:
                measurements = storage.get_all_physical_measurements(user_id)
                if len([m for m in measurements if m.height_cm is not None]) >= 2:
                    phv_result = calculate_phv(measurements, settings.date_of_birth)
                    if phv_result:
                        settings.phv_date = phv_result.get('phv_date')
                        settings.phv_age = phv_result.get('phv_age')
                        storage.save_settings(settings, user_id)
            except Exception as phv_error:
                # PHV calculation failed - not critical, just log it
                print(f"PHV auto-calculation failed: {phv_error}")
        
        return jsonify({'success': True})
        
    except (ValueError, TypeError, KeyError) as e:
        return jsonify({'success': False, 'errors': [f'Invalid match data: {str(e)}']}), 400
    except Exception as e:
        return jsonify({'success': False, 'errors': [str(e)]}), 400


@bp.route('/api/upload-photo', methods=['POST'])
@login_required
def upload_photo():
    """Upload player photo"""
    try:
        user_id = current_user.id
        if 'photo' not in request.files:
            return jsonify({'success': False, 'errors': ['No file provided']}), 400
        
        file = request.files['photo']
        if file.filename == '':
            return jsonify({'success': False, 'errors': ['No file selected']}), 400
        
        # Security: Validate file type by extension
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
        filename = secure_filename(file.filename)
        if '.' not in filename or filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
            return jsonify({'success': False, 'errors': ['Invalid file type. Allowed: PNG, JPG, JPEG, GIF, WEBP']}), 400
        
        # Security: Read file content for MIME type validation
        file_content = file.read()
        file.seek(0)  # Reset for save
        
        # Security: Validate MIME type if python-magic is available
        if MAGIC_AVAILABLE:
            mime_type = magic.from_buffer(file_content, mime=True)
            allowed_mimes = {'image/png', 'image/jpeg', 'image/gif', 'image/webp'}
            if mime_type not in allowed_mimes:
                return jsonify({'success': False, 'errors': ['Invalid file type. File content does not match extension.']}), 400
        
        # Security: Validate image can be opened and verified
        if PIL_AVAILABLE:
            try:
                img = Image.open(file)
                img.verify()
                file.seek(0)  # Reset again for save
            except Exception:
                return jsonify({'success': False, 'errors': ['Invalid image file. File may be corrupted.']}), 400
        
        # Create photos directory if it doesn't exist
        photos_dir = os.path.join(current_app.root_path, '..', 'data', 'photos')
        os.makedirs(photos_dir, exist_ok=True)
        
        # Security: Generate cryptographically random filename to prevent enumeration
        ext = filename.rsplit('.', 1)[1].lower()
        random_token = secrets.token_hex(16)
        new_filename = f"player_photo_{random_token}.{ext}"
        filepath = os.path.join(photos_dir, new_filename)
        
        # Security: Ensure filepath is within photos directory (prevent path traversal)
        photos_dir_abs = os.path.abspath(photos_dir)
        filepath_abs = os.path.abspath(filepath)
        if not filepath_abs.startswith(photos_dir_abs):
            return jsonify({'success': False, 'errors': ['Invalid file path']}), 400
        
        # Save file
        file.save(filepath)
        
        # Update settings with relative path
        settings = storage.load_settings(user_id)
        # Store relative path from project root
        relative_path = os.path.join('data', 'photos', new_filename)
        settings.player_photo_path = relative_path
        storage.save_settings(settings, user_id)
        
        return jsonify({
            'success': True,
            'photo_path': relative_path,
            'message': 'Photo uploaded successfully'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'errors': [str(e)]}), 500


@bp.route('/data/photos/<filename>')
@login_required
def serve_photo(filename):
    """Serve player photos securely"""
    try:
        # Security: Validate filename to prevent path traversal
        filename = secure_filename(filename)
        if not filename:
            return jsonify({'success': False, 'errors': ['Invalid filename']}), 400
        
        # Security: Only allow image file extensions
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
        if '.' not in filename or filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
            return jsonify({'success': False, 'errors': ['Invalid file type']}), 400
        
        # Construct file path
        photos_dir = os.path.join(current_app.root_path, '..', 'data', 'photos')
        filepath = os.path.join(photos_dir, filename)
        
        # Security: Ensure filepath is within photos directory (prevent path traversal)
        photos_dir_abs = os.path.abspath(photos_dir)
        filepath_abs = os.path.abspath(filepath)
        if not filepath_abs.startswith(photos_dir_abs):
            return jsonify({'success': False, 'errors': ['Invalid file path']}), 400
        
        # Check if file exists
        if not os.path.exists(filepath):
            # Return a placeholder image or 404
            current_app.logger.warning(f"Photo not found: {filename}")
            # Return 404 with proper content type
            from flask import Response
            return Response('Photo not found', status=404, mimetype='text/plain')
        
        # Get user's photo path from settings to verify ownership (optional check)
        user_id = current_user.id
        settings = storage.load_settings(user_id)
        
        # Check if this photo belongs to the current user (optional - allow if file exists)
        if settings.player_photo_path and filename not in settings.player_photo_path:
            # Also check if user is admin (can view any photo)
            admin_username = os.environ.get('ADMIN_USERNAME', '').strip()
            if current_user.username != admin_username:
                # Still allow if file exists - might be from import
                pass
        
        # Determine MIME type from extension
        ext = filename.rsplit('.', 1)[1].lower()
        mime_types = {
            'jpg': 'image/jpeg',
            'jpeg': 'image/jpeg',
            'png': 'image/png',
            'gif': 'image/gif',
            'webp': 'image/webp'
        }
        mimetype = mime_types.get(ext, 'image/jpeg')
        
        # Serve file
        return send_file(filepath, mimetype=mimetype)
        
    except Exception as e:
        current_app.logger.error(f"Error serving photo: {e}", exc_info=True)
        from flask import Response
        return Response('Error serving photo', status=500, mimetype='text/plain')


@bp.route('/stats')
@login_required
def get_stats():
    """Get season statistics with optional time period filter"""
    try:
        user_id = current_user.id
        period = request.args.get('period', 'all_time')  # Default to all_time
        
        # Validate period
        valid_periods = ['all_time', 'season', '12_months', '6_months', '3_months', 'last_month']
        if period not in valid_periods:
            period = 'all_time'
        
        season_stats = storage.get_season_stats(user_id, period=period)
        
        # For category stats, we need to filter matches first
        from .utils import filter_matches_by_period
        settings = storage.load_settings(user_id)
        all_matches = storage.get_all_matches(user_id)
        completed_matches = [m for m in all_matches if not m.is_fixture]
        
        if period and period != 'all_time':
            filtered_matches = filter_matches_by_period(completed_matches, period, settings.season_year)
        else:
            filtered_matches = completed_matches
        
        pre_season_matches = [m for m in filtered_matches if m.category.value == "Pre-Season Friendly"]
        league_matches = [m for m in filtered_matches if m.category.value == "League"]
        
        pre_season_stats = storage.get_category_stats("Pre-Season Friendly", user_id)
        league_stats = storage.get_category_stats("League", user_id)
        
        # Recalculate category stats from filtered matches
        def calc_stats_from_matches(match_list):
            return {
                "matches": len(match_list),
                "goals": sum(m.brodie_goals for m in match_list),
                "assists": sum(m.brodie_assists for m in match_list),
                "minutes": sum(m.minutes_played for m in match_list)
            }
        
        pre_season_stats = calc_stats_from_matches(pre_season_matches)
        league_stats = calc_stats_from_matches(league_matches)
        
        return jsonify({
            'success': True,
            'stats': {
                'season': season_stats,
                'pre_season': pre_season_stats,
                'league': league_stats
            },
            'period': period
        })
        
    except Exception as e:
        return jsonify({'success': False, 'errors': [str(e)]}), 500


@bp.route('/api/physical-measurements')
@login_required
def get_physical_measurements():
    """Get all physical measurements"""
    try:
        user_id = current_user.id
        measurements = storage.get_all_physical_measurements(user_id)
        return jsonify({
            'success': True,
            'measurements': [m.model_dump() for m in measurements]
        })
    except Exception as e:
        return jsonify({'success': False, 'errors': [str(e)]}), 500


@bp.route('/api/physical-measurements', methods=['POST'])
@login_required
def create_physical_measurement():
    """Create a new physical measurement"""
    try:
        user_id = current_user.id
        
        # Check subscription and limits
        has_access, limit, current_count, error_msg = check_subscription_and_limit(user_id, 'physical_measurements')
        if not has_access:
            return jsonify({'success': False, 'errors': [error_msg], 'limit_reached': True, 'current_count': current_count, 'limit': limit}), 403
        
        data = request.get_json()
        
        # Validate required fields
        if not data.get('date'):
            return jsonify({'success': False, 'errors': ['Date is required']}), 400
        
        if not data.get('height_cm') and not data.get('weight_kg'):
            return jsonify({'success': False, 'errors': ['At least height or weight is required']}), 400
        
        # Parse date if it's in HTML format (YYYY-MM-DD)
        date_str = str(data['date']).strip()
        if '-' in date_str and len(date_str) == 10 and date_str.count('-') == 2:
            try:
                data['date'] = parse_input_date(date_str)
            except Exception as e:
                return jsonify({'success': False, 'errors': [f'Invalid date format: {str(e)}']}), 400
        
        # Create measurement object
        measurement = PhysicalMeasurement(
            date=data['date'],
            height_cm=float(data['height_cm']) if data.get('height_cm') is not None else None,
            include_in_report=data.get('include_in_report', True),  # Default to True
            weight_kg=float(data['weight_kg']) if data.get('weight_kg') is not None else None,
            notes=data.get('notes')
        )
        
        # Save measurement
        measurement_id = storage.save_physical_measurement(measurement, user_id)
        
        # Calculate PHV if possible
        settings = storage.load_settings(user_id)
        measurements = storage.get_all_physical_measurements(user_id)
        phv_result = None
        if len(measurements) >= 2 and settings.date_of_birth:
            phv_result = calculate_phv(measurements, settings.date_of_birth)
            
            # Update settings with calculated PHV if available
            if phv_result:
                settings.phv_date = phv_result.get('phv_date')
                settings.phv_age = phv_result.get('phv_age')
                storage.save_settings(settings, user_id)
        
        return jsonify({
            'success': True,
            'measurement_id': measurement_id,
            'measurement': measurement.model_dump(),  # Return the saved measurement with converted date
            'phv_calculated': phv_result is not None,
            'phv': phv_result
        })
        
    except (ValueError, TypeError, KeyError) as e:
        return jsonify({'success': False, 'errors': [f'Invalid match data: {str(e)}']}), 400
    except Exception as e:
        return jsonify({'success': False, 'errors': [str(e)]}), 400


@bp.route('/api/physical-measurements/<measurement_id>', methods=['PUT'])
@login_required
def update_physical_measurement(measurement_id):
    """Update an existing physical measurement"""
    try:
        user_id = current_user.id
        # Get existing measurement
        existing_measurement = storage.get_physical_measurement(measurement_id, user_id)
        if not existing_measurement:
            return jsonify({'success': False, 'errors': ['Measurement not found']}), 404
        
        data = request.get_json()
        
        # Parse date if it's in HTML format
        if 'date' in data and '-' in str(data['date']) and len(str(data['date'])) == 10:
            data['date'] = parse_input_date(data['date'])
        
        # Update measurement object
        # Handle include_in_report - default to True if not set in existing measurement
        include_in_report = data.get('include_in_report')
        if include_in_report is None:
            include_in_report = getattr(existing_measurement, 'include_in_report', True)
        
        updated_measurement = PhysicalMeasurement(
            id=measurement_id,
            date=data.get('date', existing_measurement.date),
            height_cm=float(data['height_cm']) if data.get('height_cm') else existing_measurement.height_cm,
            weight_kg=float(data['weight_kg']) if data.get('weight_kg') else existing_measurement.weight_kg,
            notes=data.get('notes', existing_measurement.notes),
            include_in_report=include_in_report
        )
        
        # Save updated measurement
        storage.save_physical_measurement(updated_measurement, user_id)
        
        # Recalculate PHV if possible
        settings = storage.load_settings(user_id)
        measurements = storage.get_all_physical_measurements(user_id)
        phv_result = None
        if len(measurements) >= 2 and settings.date_of_birth:
            phv_result = calculate_phv(measurements, settings.date_of_birth)
            
            # Update settings with calculated PHV if available
            if phv_result:
                settings.phv_date = phv_result.get('phv_date')
                settings.phv_age = phv_result.get('phv_age')
                storage.save_settings(settings, user_id)
        
        return jsonify({
            'success': True,
            'measurement': updated_measurement.model_dump(),  # Return the updated measurement with converted date
            'phv_calculated': phv_result is not None,
            'phv': phv_result
        })
        
    except (ValueError, TypeError, KeyError) as e:
        return jsonify({'success': False, 'errors': [f'Invalid match data: {str(e)}']}), 400
    except Exception as e:
        return jsonify({'success': False, 'errors': [str(e)]}), 400


@bp.route('/api/physical-measurements/<measurement_id>', methods=['DELETE'])
@login_required
def delete_physical_measurement(measurement_id):
    """Delete a physical measurement"""
    user_id = current_user.id
    success = storage.delete_physical_measurement(measurement_id, user_id)
    if success:
        # Recalculate PHV after deletion
        settings = storage.load_settings(user_id)
        measurements = storage.get_all_physical_measurements(user_id)
        phv_result = None
        if len(measurements) >= 2 and settings.date_of_birth:
            phv_result = calculate_phv(measurements, settings.date_of_birth)
            
            # Update settings with calculated PHV if available
            if phv_result:
                settings.phv_date = phv_result.get('phv_date')
                settings.phv_age = phv_result.get('phv_age')
            else:
                # Clear PHV if not enough data
                settings.phv_date = None
                settings.phv_age = None
            storage.save_settings(settings, user_id)
        
        return jsonify({
            'success': True,
            'phv_calculated': phv_result is not None,
            'phv': phv_result
        })
    else:
        return jsonify({'success': False, 'errors': ['Measurement not found']}), 404


@bp.route('/api/phv/calculate')
@login_required
def calculate_phv_endpoint():
    """Calculate PHV from all measurements"""
    try:
        user_id = current_user.id
        settings = storage.load_settings(user_id)
        measurements = storage.get_all_physical_measurements(user_id)
        
        if not settings.date_of_birth:
            return jsonify({
                'success': False,
                'errors': ['Please add your date of birth in Settings to calculate PHV']
            }), 400
        
        # Validate measurements
        validation = validate_measurements_for_phv(measurements)
        
        if not validation.get('valid', False):
            return jsonify({
                'success': False,
                'phv': None,
                'validation': validation,
                'errors': [validation.get('message', 'Invalid measurements for PHV calculation')]
            })
        
        # Calculate PHV
        phv_result = calculate_phv(measurements, settings.date_of_birth)
        
        if phv_result:
            # Update settings with calculated PHV
            settings.phv_date = phv_result.get('phv_date')
            settings.phv_age = phv_result.get('phv_age')
            storage.save_settings(settings, user_id)
        elif not phv_result and len([m for m in measurements if m.height_cm is not None]) >= 2:
            # Have enough measurements but calculation failed - likely date format issue
            return jsonify({
                'success': False,
                'phv': None,
                'validation': validation,
                'errors': ['Unable to calculate PHV. Please check that all measurement dates are valid and measurements are sorted correctly.']
            })
        
        return jsonify({
            'success': True,
            'phv': phv_result,
            'validation': validation
        })
        
    except Exception as e:
        import traceback
        error_details = str(e)
        traceback.print_exc()  # Print to server logs for debugging
        return jsonify({
            'success': False, 
            'errors': [f'Error calculating PHV: {error_details}']
        }), 500


@bp.route('/player-profile')
@login_required
def player_profile():
    """Player Profile page - data loaded client-side"""
    # Redirect admin users to admin page
    admin_username = os.environ.get('ADMIN_USERNAME', '').strip()
    if admin_username and current_user.username == admin_username:
        return redirect(url_for('main.admin_users'))
    
    # Return empty/default data - actual data will be loaded client-side
    default_settings = {
        'club_name': '',
        'player_name': '',
        'season_year': ''
    }
    return render_template('player_profile.html', settings=default_settings)


@bp.route('/physical-data')
@login_required
def physical_data():
    """Physical Data page - data loaded client-side"""
    # Redirect admin users to admin page
    admin_username = os.environ.get('ADMIN_USERNAME', '').strip()
    if admin_username and current_user.username == admin_username:
        return redirect(url_for('main.admin_users'))
    
    # Return empty/default data - actual data will be loaded client-side
    default_settings = {
        'club_name': '',
        'player_name': '',
        'season_year': ''
    }
    return render_template('physical_data.html', settings=default_settings, measurements=[])


@bp.route('/api/physical-data/analysis', methods=['POST'])
@login_required
def physical_data_analysis():
    """Get comprehensive physical data analysis including PHV, predicted height, and elite comparisons - accepts data from client"""
    try:
        user_id = current_user.id
        
        # Get settings from server (most up-to-date)
        settings = storage.load_settings(user_id)
        
        # Get data from request (client-side storage for measurements and metrics)
        data = request.get_json() if request.is_json else {}
        measurements_data = data.get('measurements', [])
        physical_metrics_data = data.get('physical_metrics', [])
        
        # If client provided settings and server settings don't have date_of_birth, try client settings
        # This handles the case where settings were just saved but not yet synced
        client_settings_data = data.get('settings', {})
        if not settings.date_of_birth and client_settings_data.get('date_of_birth'):
            # Merge client settings with server settings, prioritizing client for date_of_birth
            settings_dict = settings.model_dump()
            settings_dict['date_of_birth'] = client_settings_data.get('date_of_birth')
            settings = AppSettings(**settings_dict)
        
        # Helper function to clean numeric values (convert empty strings to None)
        def clean_numeric_value(value):
            if value is None or value == '' or value == 'None':
                return None
            try:
                if isinstance(value, str):
                    value = value.strip()
                    if value == '':
                        return None
                return float(value) if '.' in str(value) else int(value)
            except (ValueError, TypeError):
                return None
        
        # Clean physical metrics data before validation
        cleaned_metrics_data = []
        for pm in physical_metrics_data:
            cleaned_pm = {}
            for key, value in pm.items():
                # Handle numeric fields
                if key in ['sprint_speed_ms', 'sprint_speed_kmh', 'sprint_10m_sec', 'sprint_20m_sec', 
                          'sprint_30m_sec', 'vertical_jump_cm', 'standing_long_jump_cm', 
                          'countermovement_jump_cm', 'agility_time_sec', 'yo_yo_test_level', 
                          'beep_test_level', 'bench_press_kg', 'squat_kg', 'deadlift_kg', 
                          'vo2_max', 'sit_and_reach_cm']:
                    cleaned_pm[key] = clean_numeric_value(value)
                elif key in ['max_heart_rate', 'resting_heart_rate']:
                    # Integer fields
                    cleaned_value = clean_numeric_value(value)
                    cleaned_pm[key] = int(cleaned_value) if cleaned_value is not None else None
                else:
                    # Non-numeric fields (id, date, notes, etc.)
                    cleaned_pm[key] = value if value else None
            cleaned_metrics_data.append(cleaned_pm)
        
        # Convert to model objects
        from .models import AppSettings, PhysicalMeasurement, PhysicalMetrics
        
        # Settings already loaded from server above
        
        # Convert measurement dates from YYYY-MM-DD to dd MMM yyyy format if needed
        cleaned_measurements_data = []
        for m in measurements_data:
            cleaned_m = m.copy()
            # Check if date is in YYYY-MM-DD format and convert
            if 'date' in cleaned_m and cleaned_m['date']:
                date_str = str(cleaned_m['date']).strip()
                if '-' in date_str and len(date_str) == 10 and date_str.count('-') == 2:
                    try:
                        # Convert from YYYY-MM-DD to dd MMM yyyy
                        cleaned_m['date'] = parse_input_date(date_str)
                    except Exception as e:
                        # If conversion fails, skip this measurement or use original
                        print(f"Warning: Could not convert date {date_str}: {e}")
                        continue
            cleaned_measurements_data.append(cleaned_m)
        
        measurements = [PhysicalMeasurement(**m) for m in cleaned_measurements_data]
        physical_metrics = [PhysicalMetrics(**pm) for pm in cleaned_metrics_data]
        
        if not settings.date_of_birth:
            return jsonify({
                'success': False,
                'errors': ['Please add your date of birth in Settings to view physical data analysis']
            }), 400
        
        # Calculate current age
        if measurements:
            valid_measurements = [m for m in measurements if m.height_cm is not None]
            if valid_measurements:
                latest_measurement = max(valid_measurements, key=lambda m: datetime.strptime(m.date, "%d %b %Y"))
                current_age = calculate_age_at_date(settings.date_of_birth, latest_measurement.date)
            else:
                current_age = calculate_age_at_date(settings.date_of_birth, datetime.now().strftime("%d %b %Y"))
        else:
            current_age = calculate_age_at_date(settings.date_of_birth, datetime.now().strftime("%d %b %Y"))
        
        # Calculate PHV
        phv_result = calculate_phv(measurements, settings.date_of_birth)
        
        # Calculate predicted adult height
        predicted_height = calculate_predicted_adult_height(
            measurements,
            settings.date_of_birth,
            current_age,
            phv_result
        )
        
        # Get elite benchmarks
        benchmarks = get_elite_benchmarks_for_age(current_age)
        
        # Compare player metrics to elite benchmarks
        comparisons = {}
        
        # Height comparison - use settings (current snapshot) or latest measurement
        height_for_comparison = settings.height_cm
        if not height_for_comparison and measurements:
            valid_measurements = [m for m in measurements if m.height_cm is not None]
            if valid_measurements:
                latest_measurement = max(valid_measurements, key=lambda m: datetime.strptime(m.date, "%d %b %Y"))
                height_for_comparison = latest_measurement.height_cm
        
        if height_for_comparison:
            comparisons['height'] = compare_to_elite(
                height_for_comparison,
                benchmarks['metrics']['height'],
                'higher_is_better'
            )
        
        # Get latest physical metrics for performance comparisons
        latest_metric = None
        if physical_metrics:
            # Sort by date and get the most recent one
            sorted_metrics = sorted(physical_metrics, key=lambda x: datetime.strptime(x.date, "%d %b %Y"), reverse=True)
            latest_metric = sorted_metrics[0] if sorted_metrics else None
        
        # Speed comparison - use latest physical metric
        if latest_metric:
            if latest_metric.sprint_speed_ms:
                comparisons['speed_ms'] = compare_to_elite(
                    latest_metric.sprint_speed_ms,
                    benchmarks['metrics']['speed'],
                    'higher_is_better'
                )
            elif latest_metric.sprint_speed_kmh:
                # Convert km/h to m/s for comparison
                speed_ms = latest_metric.sprint_speed_kmh / 3.6
                comparisons['speed_ms'] = compare_to_elite(
                    speed_ms,
                    benchmarks['metrics']['speed'],
                    'higher_is_better'
                )
                comparisons['speed_kmh'] = compare_to_elite(
                    latest_metric.sprint_speed_kmh,
                    benchmarks['metrics']['speed'],
                    'higher_is_better'
                )
        
        # Vertical jump comparison - use latest physical metric
        if latest_metric and latest_metric.vertical_jump_cm:
            comparisons['vertical_jump'] = compare_to_elite(
                latest_metric.vertical_jump_cm,
                benchmarks['metrics']['vertical_jump'],
                'higher_is_better'
            )
        
        # Agility comparison - use latest physical metric
        if latest_metric and latest_metric.agility_time_sec:
            comparisons['agility'] = compare_to_elite(
                latest_metric.agility_time_sec,
                benchmarks['metrics']['agility'],
                'lower_is_better'
            )
        
        # BMI calculation and comparison - use settings or latest measurement
        bmi = None
        weight_for_bmi = settings.weight_kg
        if not weight_for_bmi and measurements:
            valid_measurements = [m for m in measurements if m.weight_kg is not None]
            if valid_measurements:
                latest_measurement = max(valid_measurements, key=lambda m: datetime.strptime(m.date, "%d %b %Y"))
                weight_for_bmi = latest_measurement.weight_kg
        
        if height_for_comparison and weight_for_bmi:
            bmi = weight_for_bmi / ((height_for_comparison / 100) ** 2)
            bmi_benchmark = benchmarks['metrics']['body_composition']
            if bmi:
                comparisons['bmi'] = {
                    'player_value': round(bmi, 1),
                    'optimal_range': f"{bmi_benchmark['optimal_bmi_min']:.1f} - {bmi_benchmark['optimal_bmi_max']:.1f}",
                    'elite_range': f"{bmi_benchmark['elite_bmi_min']:.1f} - {bmi_benchmark['elite_bmi_max']:.1f}",
                    'in_optimal_range': bmi_benchmark['optimal_bmi_min'] <= bmi <= bmi_benchmark['optimal_bmi_max'],
                    'in_elite_range': bmi_benchmark['elite_bmi_min'] <= bmi <= bmi_benchmark['elite_bmi_max'],
                    'description': bmi_benchmark['description']
                }
        
        return jsonify({
            'success': True,
            'current_age': current_age,
            'phv': phv_result,
            'predicted_height': predicted_height,
            'benchmarks': benchmarks,
            'comparisons': comparisons,
            'bmi': bmi
        })
        
    except Exception as e:
        import traceback
        error_details = str(e)
        traceback.print_exc()
        return jsonify({
            'success': False,
            'errors': ['Unable to analyze physical data at this time. Please try again or check your data.']
        }), 500


@bp.route('/achievements')
@login_required
def achievements():
    """Achievements page - data loaded client-side"""
    # Redirect admin users to admin page
    admin_username = os.environ.get('ADMIN_USERNAME', '').strip()
    if admin_username and current_user.username == admin_username:
        return redirect(url_for('main.admin_users'))
    
    # Return empty/default data - actual data will be loaded client-side
    default_settings = {
        'club_name': '',
        'player_name': '',
        'season_year': ''
    }
    return render_template('achievements.html', settings=default_settings, achievements=[])


@bp.route('/api/achievements')
@login_required
def get_achievements():
    """Get all achievements"""
    try:
        user_id = current_user.id
        achievements_list = storage.get_all_achievements(user_id)
        # Sort by date descending
        achievements_list.sort(key=lambda x: datetime.strptime(x.date, "%d %b %Y"), reverse=True)
        return jsonify({
            'success': True,
            'achievements': [a.model_dump() for a in achievements_list]
        })
    except Exception as e:
        return jsonify({'success': False, 'errors': [str(e)]}), 500


@bp.route('/api/achievements', methods=['POST'])
@login_required
def create_achievement():
    """Create a new achievement"""
    try:
        user_id = current_user.id
        
        # Check subscription and limits
        has_access, limit, current_count, error_msg = check_subscription_and_limit(user_id, 'achievements')
        if not has_access:
            return jsonify({'success': False, 'errors': [error_msg], 'limit_reached': True, 'current_count': current_count, 'limit': limit}), 403
        
        data = request.get_json()
        
        # Validate required fields
        if not data.get('title'):
            return jsonify({'success': False, 'errors': ['Title is required']}), 400
        
        if not data.get('date'):
            return jsonify({'success': False, 'errors': ['Date is required']}), 400
        
        if not data.get('category'):
            return jsonify({'success': False, 'errors': ['Category is required']}), 400
        
        # Parse date if it's in HTML format (YYYY-MM-DD)
        date_str = str(data['date']).strip()
        if '-' in date_str and len(date_str) == 10 and date_str.count('-') == 2:
            try:
                data['date'] = parse_input_date(date_str)
            except Exception as e:
                return jsonify({'success': False, 'errors': [f'Invalid date format: {str(e)}']}), 400
        
        # Handle position-specific numeric fields
        goals = None
        if 'goals' in data and data['goals'] is not None and data['goals'] != '':
            try:
                goals = int(data['goals'])
            except (ValueError, TypeError):
                goals = None
        
        clean_sheets = None
        if 'clean_sheets' in data and data['clean_sheets'] is not None and data['clean_sheets'] != '':
            try:
                clean_sheets = int(data['clean_sheets'])
            except (ValueError, TypeError):
                clean_sheets = None
        
        # Create achievement object
        achievement = Achievement(
            title=data['title'].strip(),
            category=data['category'].strip(),
            date=data['date'],
            season=data.get('season'),
            description=data.get('description'),
            notes=data.get('notes'),
            goals=goals,
            clean_sheets=clean_sheets
        )
        
        # Save achievement
        achievement_id = storage.save_achievement(achievement, user_id)
        
        return jsonify({
            'success': True,
            'achievement_id': achievement_id
        })
        
    except (ValueError, TypeError, KeyError) as e:
        return jsonify({'success': False, 'errors': [f'Invalid achievement data: {str(e)}']}), 400
    except Exception as e:
        return jsonify({'success': False, 'errors': [str(e)]}), 400


@bp.route('/api/achievements/<achievement_id>', methods=['PUT'])
@login_required
def update_achievement(achievement_id):
    """Update an existing achievement"""
    try:
        user_id = current_user.id
        # Get existing achievement
        existing_achievement = storage.get_achievement(achievement_id, user_id)
        if not existing_achievement:
            return jsonify({'success': False, 'errors': ['Achievement not found']}), 404
        
        data = request.get_json()
        
        # Parse date if it's in HTML format
        if 'date' in data and '-' in str(data['date']) and len(str(data['date'])) == 10:
            data['date'] = parse_input_date(data['date'])
        
        # Handle position-specific numeric fields
        goals = existing_achievement.goals
        if 'goals' in data:
            if data['goals'] is None or data['goals'] == '':
                goals = None
            else:
                try:
                    goals = int(data['goals'])
                except (ValueError, TypeError):
                    goals = None
        
        clean_sheets = existing_achievement.clean_sheets
        if 'clean_sheets' in data:
            if data['clean_sheets'] is None or data['clean_sheets'] == '':
                clean_sheets = None
            else:
                try:
                    clean_sheets = int(data['clean_sheets'])
                except (ValueError, TypeError):
                    clean_sheets = None
        
        # Update achievement object
        updated_achievement = Achievement(
            id=achievement_id,
            title=data.get('title', existing_achievement.title).strip(),
            category=data.get('category', existing_achievement.category).strip(),
            date=data.get('date', existing_achievement.date),
            season=data.get('season', existing_achievement.season),
            description=data.get('description', existing_achievement.description),
            notes=data.get('notes', existing_achievement.notes),
            goals=goals,
            clean_sheets=clean_sheets
        )
        
        # Save updated achievement
        storage.save_achievement(updated_achievement, user_id)
        
        return jsonify({
            'success': True
        })
        
    except (ValueError, TypeError, KeyError) as e:
        return jsonify({'success': False, 'errors': [f'Invalid achievement data: {str(e)}']}), 400
    except Exception as e:
        return jsonify({'success': False, 'errors': [str(e)]}), 400


@bp.route('/api/achievements/<achievement_id>', methods=['DELETE'])
@login_required
def delete_achievement(achievement_id):
    """Delete an achievement"""
    user_id = current_user.id
    success = storage.delete_achievement(achievement_id, user_id)
    if success:
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'errors': ['Achievement not found']}), 404


@bp.route('/scout-pdf', methods=['POST'])
@login_required
def generate_scout_pdf_route():
    """Generate scout-friendly PDF report - accepts data from client"""
    try:
        # Check subscription status - try multiple methods
        subscription = storage.get_subscription_by_user_id(current_user.id)
        
        # If no subscription found by user_id, check if there's exactly one active subscription
        # This handles cases where user_id format changed (e.g., client-side ID vs server-side ID)
        if not subscription:
            subscriptions = storage.load_subscriptions()
            active_subs = [s for s in subscriptions if s.get('status', '').lower() == 'active']
            
            # If there's exactly one active subscription, it's likely for the current user
            # Update it to use the correct user_id
            if len(active_subs) == 1:
                sub_data = active_subs[0]
                try:
                    # Create subscription object and update user_id
                    if 'status' in sub_data and isinstance(sub_data['status'], str):
                        sub_data['status'] = SubscriptionStatus(sub_data['status'].lower())
                    subscription = Subscription(**sub_data)
                    subscription.user_id = current_user.id
                    storage.save_subscription(subscription)
                    current_app.logger.info(f"Updated subscription user_id from {sub_data.get('user_id')} to {current_user.id}")
                except Exception as e:
                    current_app.logger.error(f"Error updating subscription: {e}")
                    subscription = None
        
        # Debug logging
        current_app.logger.info(f"Scout PDF generation request for user {current_user.id}")
        current_app.logger.info(f"Subscription found: {subscription is not None}")
        if subscription:
            current_app.logger.info(f"Subscription status: {subscription.status}, type: {type(subscription.status)}")
            current_app.logger.info(f"Subscription status == ACTIVE: {subscription.status == SubscriptionStatus.ACTIVE}")
            current_app.logger.info(f"Subscription status value: {subscription.status.value if hasattr(subscription.status, 'value') else subscription.status}")
        
        # Check if subscription is active - simplified check
        is_active = False
        if subscription:
            # Get status value (works for both enum and string)
            status_value = subscription.status.value if hasattr(subscription.status, 'value') else str(subscription.status)
            is_active = status_value.lower() == SubscriptionStatus.ACTIVE.value.lower()
            current_app.logger.info(f"Status value: {status_value}, ACTIVE value: {SubscriptionStatus.ACTIVE.value}, is_active: {is_active}")
        
        if not subscription or not is_active:
            current_app.logger.warning(f"Scout PDF generation blocked for user {current_user.id}: subscription={subscription is not None}, is_active={is_active}")
            return jsonify({
                'success': False,
                'errors': ['Scout report generation is a premium feature. Please subscribe to unlock this feature.'],
                'debug': {
                    'has_subscription': subscription is not None,
                    'status': subscription.status if subscription else None,
                    'user_id': current_user.id
                }
            }), 403
        
        data = request.get_json() if request.is_json else {}
        user_id = current_user.id
        
        # Get data from request (client-side storage)
        matches_data = data.get('matches', [])
        settings_data = data.get('settings', {})
        physical_measurements_data = data.get('physical_measurements', [])
        achievements_data = data.get('achievements', [])
        club_history_data = data.get('club_history', [])
        training_camps_data = data.get('training_camps', [])
        physical_metrics_data = data.get('physical_metrics', [])
        references_data = data.get('references', [])
        period = data.get('period', 'all_time')
        
        # Validate period
        valid_periods = ['all_time', 'season', '12_months', '6_months', '3_months', 'last_month']
        if period not in valid_periods:
            period = 'all_time'
        
        # Load settings from server (most up-to-date, includes player profile data)
        # Merge with client settings to ensure we have all data
        try:
            server_settings = storage.load_settings(user_id)
            if server_settings:
                # Merge server settings (source of truth) with client settings
                # Server settings take priority, but client settings can fill in gaps
                server_dict = server_settings.model_dump()
                # Only update with client settings that are not None/empty in server settings
                # AND the client value is not None/empty
                for key, value in settings_data.items():
                    if (key not in server_dict or 
                        server_dict[key] is None or 
                        server_dict[key] == '' or
                        (isinstance(server_dict[key], list) and len(server_dict[key]) == 0)):
                        # Only use client value if it's not None/empty
                        if value is not None and value != '' and not (isinstance(value, list) and len(value) == 0):
                            server_dict[key] = value
                settings_data = server_dict
        except Exception as e:
            # If loading from server fails, use client settings
            print(f"Warning: Could not load settings from server: {e}")
            pass
        
        # Convert data to model objects
        from .models import Match, AppSettings, PhysicalMeasurement, PhysicalMetrics, Achievement, ClubHistory, TrainingCamp, Reference
        
        matches = [Match(**m) for m in matches_data]
        settings = AppSettings(**settings_data) if settings_data else AppSettings()
        
        # Convert physical measurement dates from YYYY-MM-DD to dd MMM yyyy format if needed
        cleaned_measurements_data = []
        for pm in physical_measurements_data:
            if not pm or not isinstance(pm, dict):
                continue
            cleaned_pm = pm.copy()
            if 'date' in cleaned_pm and cleaned_pm['date']:
                date_str = str(cleaned_pm['date']).strip()
                if date_str:
                    # Check if date is already in correct format (dd MMM yyyy)
                    if not re.match(r'^\d{1,2} [A-Za-z]{3} \d{4}', date_str):
                        # Not in correct format, try to convert
                        # Check if it's YYYY-MM-DD format
                        if re.match(r'^\d{4}-\d{2}-\d{2}', date_str):
                            try:
                                converted_date = parse_input_date(date_str)
                                # Verify conversion worked (should be in dd MMM yyyy format)
                                if converted_date != date_str and re.match(r'^\d{1,2} [A-Za-z]{3} \d{4}', converted_date):
                                    cleaned_pm['date'] = converted_date
                                    print(f"Converted date: {date_str} -> {cleaned_pm['date']}")
                                else:
                                    print(f"Warning: Date conversion returned unexpected format: {date_str} -> {converted_date}")
                                    continue
                            except Exception as e:
                                print(f"Error: Could not convert date {date_str}: {e}")
                                continue  # Skip this measurement if date conversion fails
                        else:
                            # Try to convert anyway (might be another format)
                            try:
                                converted_date = parse_input_date(date_str)
                                if converted_date != date_str and re.match(r'^\d{1,2} [A-Za-z]{3} \d{4}', converted_date):
                                    cleaned_pm['date'] = converted_date
                                    print(f"Converted date (fallback): {date_str} -> {cleaned_pm['date']}")
                                else:
                                    print(f"Warning: Could not convert date format: {date_str}")
                                    continue
                            except Exception as e:
                                print(f"Error: Could not convert date {date_str}: {e}")
                                continue
            cleaned_measurements_data.append(cleaned_pm)
        
        print(f"Processed {len(cleaned_measurements_data)} measurements out of {len(physical_measurements_data)}")
        physical_measurements = [PhysicalMeasurement(**pm) for pm in cleaned_measurements_data]
        
        # Convert achievement dates from YYYY-MM-DD to dd MMM yyyy format if needed
        cleaned_achievements_data = []
        for a in achievements_data:
            cleaned_a = a.copy()
            if 'date' in cleaned_a and cleaned_a['date']:
                date_str = str(cleaned_a['date']).strip()
                # Check if date is in YYYY-MM-DD format and convert
                if '-' in date_str and len(date_str) == 10 and date_str.count('-') == 2:
                    try:
                        cleaned_a['date'] = parse_input_date(date_str)
                    except Exception as e:
                        print(f"Warning: Could not convert date {date_str}: {e}")
                        continue  # Skip this achievement if date conversion fails
            cleaned_achievements_data.append(cleaned_a)
        
        achievements = [Achievement(**a) for a in cleaned_achievements_data]
        club_history = [ClubHistory(**ch) for ch in club_history_data]
        training_camps = [TrainingCamp(**tc) for tc in training_camps_data]
        
        # Convert physical metrics dates from YYYY-MM-DD to dd MMM yyyy format if needed
        # Also clean numeric fields (convert empty strings to None)
        cleaned_metrics_data = []
        numeric_fields = [
            'sprint_speed_ms', 'sprint_speed_kmh', 'sprint_10m_sec', 'sprint_20m_sec', 'sprint_30m_sec',
            'vertical_jump_cm', 'standing_long_jump_cm', 'countermovement_jump_cm',
            'agility_time_sec', 'yo_yo_test_level', 'beep_test_level',
            'bench_press_kg', 'squat_kg', 'deadlift_kg',
            'vo2_max', 'sit_and_reach_cm'
        ]
        integer_fields = ['max_heart_rate', 'resting_heart_rate']
        
        for pm in physical_metrics_data:
            if not pm or not isinstance(pm, dict):
                continue
            cleaned_pm = {}
            # Initialize ALL numeric fields to None first
            for field in numeric_fields:
                cleaned_pm[field] = None
            for field in integer_fields:
                cleaned_pm[field] = None
            
            for key, value in pm.items():
                # Handle numeric fields - convert empty strings to None
                if key in numeric_fields:
                    # Check for empty string, None, or whitespace-only string
                    if value is None or value == '' or (isinstance(value, str) and value.strip() == ''):
                        cleaned_pm[key] = None
                    else:
                        try:
                            # Try to convert to float
                            cleaned_pm[key] = float(value)
                        except (ValueError, TypeError):
                            cleaned_pm[key] = None
                # Handle integer fields - convert empty strings to None
                elif key in integer_fields:
                    # Check for empty string, None, or whitespace-only string
                    if value is None or value == '' or (isinstance(value, str) and value.strip() == ''):
                        cleaned_pm[key] = None
                    else:
                        try:
                            # Try to convert to int
                            cleaned_pm[key] = int(value)
                        except (ValueError, TypeError):
                            cleaned_pm[key] = None
                # Handle date field
                elif key == 'date' and value:
                    date_str = str(value).strip()
                    if date_str:
                        # Check if date is already in correct format (dd MMM yyyy)
                        if not re.match(r'^\d{1,2} [A-Za-z]{3} \d{4}', date_str):
                            # Check if it's YYYY-MM-DD format
                            if re.match(r'^\d{4}-\d{2}-\d{2}', date_str):
                                try:
                                    converted_date = parse_input_date(date_str)
                                    if converted_date != date_str and re.match(r'^\d{1,2} [A-Za-z]{3} \d{4}', converted_date):
                                        cleaned_pm[key] = converted_date
                                    else:
                                        print(f"Warning: Date conversion failed for {date_str}")
                                        continue  # Skip this metric
                                except Exception as e:
                                    print(f"Error: Could not convert date {date_str}: {e}")
                                    continue  # Skip this metric
                            else:
                                # Try to convert anyway
                                try:
                                    converted_date = parse_input_date(date_str)
                                    if converted_date != date_str and re.match(r'^\d{1,2} [A-Za-z]{3} \d{4}', converted_date):
                                        cleaned_pm[key] = converted_date
                                    else:
                                        print(f"Warning: Could not convert date format: {date_str}")
                                        continue
                                except Exception as e:
                                    print(f"Error: Could not convert date {date_str}: {e}")
                                    continue
                        else:
                            cleaned_pm[key] = date_str
                    else:
                        continue  # Skip if date is empty
                else:
                    # Copy other fields as-is (id, notes, etc.) - but skip numeric fields
                    if key not in numeric_fields + integer_fields:
                        cleaned_pm[key] = value
            
            # CRITICAL: Ensure all numeric fields are explicitly set to None (never empty strings)
            for field in numeric_fields + integer_fields:
                # Always set to None if missing, empty string, or invalid
                if field not in cleaned_pm:
                    cleaned_pm[field] = None
                elif cleaned_pm[field] == '':
                    cleaned_pm[field] = None
                elif isinstance(cleaned_pm[field], str):
                    if not cleaned_pm[field].strip():
                        cleaned_pm[field] = None
                    else:
                        # Try to convert string to number
                        try:
                            if field in integer_fields:
                                cleaned_pm[field] = int(cleaned_pm[field])
                            else:
                                cleaned_pm[field] = float(cleaned_pm[field])
                        except (ValueError, TypeError):
                            cleaned_pm[field] = None
            
            # Only add if we have a valid date
            if 'date' in cleaned_pm and cleaned_pm['date']:
                cleaned_metrics_data.append(cleaned_pm)
        
        print(f"Processed {len(cleaned_metrics_data)} physical metrics out of {len(physical_metrics_data)}")
        
        # AGGRESSIVE final cleanup: create completely new dicts with only valid values
        final_cleaned_metrics = []
        for pm in cleaned_metrics_data:
            final_pm = {}
            # Copy required fields
            if 'id' in pm:
                final_pm['id'] = pm['id']
            if 'date' in pm and pm['date']:
                final_pm['date'] = pm['date']
            if 'notes' in pm:
                final_pm['notes'] = pm.get('notes')
            
            # Process ALL numeric fields - explicitly convert empty strings to None
            for field in numeric_fields:
                value = pm.get(field)
                if value is None:
                    final_pm[field] = None
                elif value == '':
                    final_pm[field] = None
                elif isinstance(value, str):
                    if value.strip() == '':
                        final_pm[field] = None
                    else:
                        try:
                            final_pm[field] = float(value)
                        except (ValueError, TypeError):
                            final_pm[field] = None
                elif isinstance(value, (int, float)):
                    final_pm[field] = float(value)
                else:
                    final_pm[field] = None
            
            for field in integer_fields:
                value = pm.get(field)
                if value is None:
                    final_pm[field] = None
                elif value == '':
                    final_pm[field] = None
                elif isinstance(value, str):
                    if value.strip() == '':
                        final_pm[field] = None
                    else:
                        try:
                            final_pm[field] = int(value)
                        except (ValueError, TypeError):
                            final_pm[field] = None
                elif isinstance(value, int):
                    final_pm[field] = value
                elif isinstance(value, float):
                    final_pm[field] = int(value)
                else:
                    final_pm[field] = None
            
            final_cleaned_metrics.append(final_pm)
        
        # Debug: print first metric to verify
        if final_cleaned_metrics:
            print(f"Final cleaned metric sample - keys: {list(final_cleaned_metrics[0].keys())}")
            for field in ['vertical_jump_cm', 'countermovement_jump_cm', 'agility_time_sec', 'yo_yo_test_level', 'bench_press_kg']:
                if field in final_cleaned_metrics[0]:
                    val = final_cleaned_metrics[0][field]
                    print(f"  {field}: {repr(val)} (type: {type(val).__name__})")
        
        # Create PhysicalMetrics - ONLY include fields with valid values (skip empty strings entirely)
        physical_metrics = []
        for i, pm in enumerate(final_cleaned_metrics):
            try:
                # Build dict with ONLY valid values - don't include empty strings at all
                safe_pm = {}
                
                # Required fields
                if 'id' in pm:
                    safe_pm['id'] = pm['id']
                if 'date' in pm and pm['date']:
                    safe_pm['date'] = pm['date']
                if 'notes' in pm and pm.get('notes'):
                    safe_pm['notes'] = pm['notes']
                
                # Only include numeric fields if they have valid (non-empty) values
                for field in numeric_fields:
                    val = pm.get(field)
                    # Skip if None, empty string, or whitespace-only string
                    if val is None or val == '' or (isinstance(val, str) and not val.strip()):
                        continue  # Don't include this field - Pydantic will use default None
                    # Try to convert to float
                    try:
                        safe_pm[field] = float(val)
                    except (ValueError, TypeError):
                        continue  # Skip invalid values
                
                for field in integer_fields:
                    val = pm.get(field)
                    # Skip if None, empty string, or whitespace-only string
                    if val is None or val == '' or (isinstance(val, str) and not val.strip()):
                        continue  # Don't include this field - Pydantic will use default None
                    # Try to convert to int
                    try:
                        safe_pm[field] = int(val)
                    except (ValueError, TypeError):
                        continue  # Skip invalid values
                
                # Create PhysicalMetrics - empty string fields won't be in the dict, so Pydantic uses None
                physical_metrics.append(PhysicalMetrics(**safe_pm))
            except Exception as e:
                print(f"Error creating PhysicalMetrics for metric {i}: {e}")
                print(f"  Problematic data: {pm}")
                import traceback
                traceback.print_exc()
                # Skip this metric
                continue
        references = [Reference(**r) for r in references_data]
        
        # Create output directory
        output_dir = os.path.join(current_app.root_path, '..', 'output')
        os.makedirs(output_dir, exist_ok=True)
        
        # Generate PDF with period filter
        pdf_path = generate_scout_pdf(
            matches, 
            settings, 
            achievements, 
            club_history, 
            physical_measurements,
            training_camps,
            physical_metrics,
            references,
            output_dir,
            period=period
        )
        
        # Return the PDF file directly
        return send_file(
            pdf_path,
            as_attachment=True,
            download_name=os.path.basename(pdf_path),
            mimetype='application/pdf'
        )
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error generating scout PDF: {error_details}")
        current_app.logger.error(f"Scout PDF generation error: {error_details}")
        return jsonify({'success': False, 'errors': [str(e)]}), 500


@bp.route('/club-history')
def club_history_page():
    """Club History page"""
    settings = storage.load_settings()
    club_history = storage.get_all_club_history()
    # Sort by season descending (most recent first)
    club_history.sort(key=lambda x: x.season, reverse=True)
    return render_template('club_history.html', settings=settings, club_history=club_history)


@bp.route('/api/club-history')
def get_club_history():
    """Get all club history entries"""
    try:
        club_history = storage.get_all_club_history()
        # Sort by season descending
        club_history.sort(key=lambda x: x.season, reverse=True)
        return jsonify({
            'success': True,
            'club_history': [h.model_dump() for h in club_history]
        })
    except Exception as e:
        return jsonify({'success': False, 'errors': [str(e)]}), 500


@bp.route('/api/club-history/<entry_id>')
def get_club_history_entry(entry_id):
    """Get a specific club history entry"""
    entry = storage.get_club_history_entry(entry_id)
    if entry:
        return jsonify({'success': True, 'club_history': entry.model_dump()})
    else:
        return jsonify({'success': False, 'errors': ['Club history entry not found']}), 404


@bp.route('/api/club-history', methods=['POST'])
def create_club_history():
    """Create a new club history entry"""
    try:
        data = request.get_json()
        
        # Validate required fields
        if not data.get('club_name'):
            return jsonify({'success': False, 'errors': ['Club name is required']}), 400
        
        if not data.get('season'):
            return jsonify({'success': False, 'errors': ['Season is required']}), 400
        
        # Create club history object
        club_history = ClubHistory(
            club_name=data['club_name'].strip(),
            season=data['season'].strip(),
            age_group=data.get('age_group'),
            position=data.get('position'),
            achievements=data.get('achievements'),
            notes=data.get('notes')
        )
        
        # Save club history entry
        entry_id = storage.save_club_history_entry(club_history)
        
        return jsonify({
            'success': True,
            'entry_id': entry_id
        })
        
    except (ValueError, TypeError, KeyError) as e:
        return jsonify({'success': False, 'errors': [f'Invalid club history data: {str(e)}']}), 400
    except Exception as e:
        return jsonify({'success': False, 'errors': [str(e)]}), 400


@bp.route('/api/club-history/<entry_id>', methods=['PUT'])
def update_club_history(entry_id):
    """Update an existing club history entry"""
    try:
        # Get existing entry
        existing_entry = storage.get_club_history_entry(entry_id)
        if not existing_entry:
            return jsonify({'success': False, 'errors': ['Club history entry not found']}), 404
        
        data = request.get_json()
        
        # Update club history object
        updated_entry = ClubHistory(
            id=entry_id,
            club_name=data.get('club_name', existing_entry.club_name).strip(),
            season=data.get('season', existing_entry.season).strip(),
            age_group=data.get('age_group', existing_entry.age_group),
            position=data.get('position', existing_entry.position),
            achievements=data.get('achievements', existing_entry.achievements),
            notes=data.get('notes', existing_entry.notes)
        )
        
        # Save updated entry
        storage.save_club_history_entry(updated_entry)
        
        return jsonify({
            'success': True
        })
        
    except (ValueError, TypeError, KeyError) as e:
        return jsonify({'success': False, 'errors': [f'Invalid club history data: {str(e)}']}), 400
    except Exception as e:
        return jsonify({'success': False, 'errors': [str(e)]}), 400


@bp.route('/api/club-history/<entry_id>', methods=['DELETE'])
def delete_club_history(entry_id):
    """Delete a club history entry"""
    success = storage.delete_club_history_entry(entry_id)
    if success:
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'errors': ['Club history entry not found']}), 404


@bp.route('/training-camps')
def training_camps_page():
    """Training camps management page"""
    settings = storage.load_settings()
    training_camps = storage.get_all_training_camps()
    # Sort by start date (most recent first)
    training_camps.sort(key=lambda x: datetime.strptime(x.start_date, "%d %b %Y"), reverse=True)
    return render_template('training_camps.html', settings=settings, training_camps=training_camps)


@bp.route('/api/training-camps')
def get_training_camps():
    """Get all training camp entries"""
    try:
        training_camps = storage.get_all_training_camps()
        # Sort by start date descending
        training_camps.sort(key=lambda x: datetime.strptime(x.start_date, "%d %b %Y"), reverse=True)
        return jsonify({
            'success': True,
            'training_camps': [c.model_dump() for c in training_camps]
        })
    except Exception as e:
        return jsonify({'success': False, 'errors': [str(e)]}), 500


@bp.route('/api/training-camps/<camp_id>')
def get_training_camp(camp_id):
    """Get a specific training camp entry"""
    camp = storage.get_training_camp(camp_id)
    if camp:
        return jsonify({'success': True, 'training_camp': camp.model_dump()})
    else:
        return jsonify({'success': False, 'errors': ['Training camp not found']}), 404


@bp.route('/api/training-camps', methods=['POST'])
def create_training_camp():
    """Create a new training camp entry"""
    try:
        data = request.get_json()
        
        # Validate required fields
        if not data.get('camp_name'):
            return jsonify({'success': False, 'errors': ['Camp name is required']}), 400
        
        if not data.get('organizer'):
            return jsonify({'success': False, 'errors': ['Organizer is required']}), 400
        
        if not data.get('location'):
            return jsonify({'success': False, 'errors': ['Location is required']}), 400
        
        if not data.get('start_date'):
            return jsonify({'success': False, 'errors': ['Start date is required']}), 400
        
        # Create training camp object
        training_camp = TrainingCamp(
            camp_name=data['camp_name'].strip(),
            organizer=data['organizer'].strip(),
            location=data['location'].strip(),
            start_date=data['start_date'].strip(),
            end_date=data.get('end_date', '').strip() if data.get('end_date') and data.get('end_date').strip() else None,
            duration_days=data.get('duration_days'),
            age_group=data.get('age_group'),
            focus_area=data.get('focus_area'),
            achievements=data.get('achievements'),
            notes=data.get('notes')
        )
        
        # Save training camp entry
        camp_id = storage.save_training_camp(training_camp)
        
        return jsonify({
            'success': True,
            'camp_id': camp_id
        })
        
    except (ValueError, TypeError, KeyError) as e:
        return jsonify({'success': False, 'errors': [f'Invalid training camp data: {str(e)}']}), 400
    except Exception as e:
        return jsonify({'success': False, 'errors': [str(e)]}), 400


@bp.route('/api/training-camps/<camp_id>', methods=['PUT'])
def update_training_camp(camp_id):
    """Update an existing training camp entry"""
    try:
        data = request.get_json()
        
        # Get existing camp
        existing_camp = storage.get_training_camp(camp_id)
        if not existing_camp:
            return jsonify({'success': False, 'errors': ['Training camp not found']}), 404
        
        # Update fields
        training_camp = TrainingCamp(
            id=camp_id,
            camp_name=data.get('camp_name', existing_camp.camp_name).strip(),
            organizer=data.get('organizer', existing_camp.organizer).strip(),
            location=data.get('location', existing_camp.location).strip(),
            start_date=data.get('start_date', existing_camp.start_date).strip(),
            end_date=data.get('end_date', '').strip() if data.get('end_date') and data.get('end_date').strip() else None,
            duration_days=data.get('duration_days'),
            age_group=data.get('age_group'),
            focus_area=data.get('focus_area'),
            achievements=data.get('achievements'),
            notes=data.get('notes')
        )
        
        # Save updated camp
        storage.save_training_camp(training_camp)
        
        return jsonify({'success': True})
        
    except (ValueError, TypeError, KeyError) as e:
        return jsonify({'success': False, 'errors': [f'Invalid training camp data: {str(e)}']}), 400
    except Exception as e:
        return jsonify({'success': False, 'errors': [str(e)]}), 400


@bp.route('/api/training-camps/<camp_id>', methods=['DELETE'])
def delete_training_camp(camp_id):
    """Delete a training camp entry"""
    success = storage.delete_training_camp(camp_id)
    if success:
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'errors': ['Training camp not found']}), 404


@bp.route('/physical-metrics')
@login_required
def physical_metrics_page():
    """Physical metrics management page - data loaded client-side"""
    # Redirect admin users to admin page
    admin_username = os.environ.get('ADMIN_USERNAME', '').strip()
    if admin_username and current_user.username == admin_username:
        return redirect(url_for('main.admin_users'))
    
    # Return empty/default data - actual data will be loaded client-side
    default_settings = {
        'club_name': '',
        'player_name': '',
        'season_year': ''
    }
    return render_template('physical_metrics.html', settings=default_settings, metrics=[])


@bp.route('/api/physical-metrics')
@login_required
def get_physical_metrics():
    """Get all physical metric entries"""
    try:
        user_id = current_user.id
        metrics = storage.get_all_physical_metrics(user_id)
        # Sort by date descending
        metrics.sort(key=lambda x: datetime.strptime(x.date, "%d %b %Y"), reverse=True)
        return jsonify({
            'success': True,
            'metrics': [m.model_dump() for m in metrics]
        })
    except Exception as e:
        return jsonify({'success': False, 'errors': [str(e)]}), 500


@bp.route('/api/physical-metrics/<metric_id>')
@login_required
def get_physical_metric(metric_id):
    """Get a specific physical metric entry"""
    user_id = current_user.id
    metric = storage.get_physical_metric(metric_id, user_id)
    if metric:
        return jsonify({'success': True, 'metric': metric.model_dump()})
    else:
        return jsonify({'success': False, 'errors': ['Physical metric not found']}), 404


@bp.route('/api/physical-metrics', methods=['POST'])
@login_required
def create_physical_metric():
    """Create a new physical metric entry"""
    try:
        user_id = current_user.id
        
        # Check subscription and limits
        has_access, limit, current_count, error_msg = check_subscription_and_limit(user_id, 'physical_metrics')
        if not has_access:
            return jsonify({'success': False, 'errors': [error_msg], 'limit_reached': True, 'current_count': current_count, 'limit': limit}), 403
        
        data = request.get_json()
        
        # Validate required fields
        if not data.get('date'):
            return jsonify({'success': False, 'errors': ['Date is required']}), 400
        
        # Parse date if needed
        date_str = str(data['date']).strip()
        if '-' in date_str and len(date_str) == 10 and date_str.count('-') == 2:
            try:
                date_str = parse_input_date(date_str)
            except Exception:
                pass  # Let the model validator handle it
        
        # Helper function to convert to float or None
        def to_float_or_none(value):
            if value is None or value == '':
                return None
            try:
                return float(value)
            except (ValueError, TypeError):
                return None
        
        def to_int_or_none(value):
            if value is None or value == '':
                return None
            try:
                return int(value)
            except (ValueError, TypeError):
                return None
        
        # Auto-convert sprint speed: if m/s is provided, calculate km/h; if only km/h, calculate m/s
        sprint_speed_ms = to_float_or_none(data.get('sprint_speed_ms'))
        sprint_speed_kmh = to_float_or_none(data.get('sprint_speed_kmh'))
        
        if sprint_speed_ms and not sprint_speed_kmh:
            # Calculate km/h from m/s
            sprint_speed_kmh = sprint_speed_ms * 3.6
        elif sprint_speed_kmh and not sprint_speed_ms:
            # Calculate m/s from km/h
            sprint_speed_ms = sprint_speed_kmh / 3.6
        elif sprint_speed_ms and sprint_speed_kmh:
            # Both provided - prefer m/s and recalculate km/h for consistency
            sprint_speed_kmh = sprint_speed_ms * 3.6
        
        # Create physical metric object
        metric = PhysicalMetrics(
            date=date_str,
            sprint_speed_ms=sprint_speed_ms,
            sprint_speed_kmh=sprint_speed_kmh,
            sprint_10m_sec=to_float_or_none(data.get('sprint_10m_sec')),
            sprint_20m_sec=to_float_or_none(data.get('sprint_20m_sec')),
            sprint_30m_sec=to_float_or_none(data.get('sprint_30m_sec')),
            vertical_jump_cm=to_float_or_none(data.get('vertical_jump_cm')),
            standing_long_jump_cm=to_float_or_none(data.get('standing_long_jump_cm')),
            countermovement_jump_cm=to_float_or_none(data.get('countermovement_jump_cm')),
            agility_time_sec=to_float_or_none(data.get('agility_time_sec')),
            yo_yo_test_level=to_float_or_none(data.get('yo_yo_test_level')),
            beep_test_level=to_float_or_none(data.get('beep_test_level')),
            bench_press_kg=to_float_or_none(data.get('bench_press_kg')),
            squat_kg=to_float_or_none(data.get('squat_kg')),
            deadlift_kg=to_float_or_none(data.get('deadlift_kg')),
            vo2_max=to_float_or_none(data.get('vo2_max')),
            max_heart_rate=to_int_or_none(data.get('max_heart_rate')),
            resting_heart_rate=to_int_or_none(data.get('resting_heart_rate')),
            sit_and_reach_cm=to_float_or_none(data.get('sit_and_reach_cm')),
            notes=data.get('notes'),
            include_in_report=data.get('include_in_report', True)  # Default to True
        )
        
        # Save metric entry
        metric_id = storage.save_physical_metric(metric, user_id)
        
        return jsonify({
            'success': True,
            'metric_id': metric_id
        })
        
    except (ValueError, TypeError, KeyError) as e:
        return jsonify({'success': False, 'errors': [f'Invalid physical metric data: {str(e)}']}), 400
    except Exception as e:
        return jsonify({'success': False, 'errors': [str(e)]}), 400


@bp.route('/api/physical-metrics/<metric_id>', methods=['PUT'])
@login_required
def update_physical_metric(metric_id):
    """Update an existing physical metric entry"""
    try:
        user_id = current_user.id
        data = request.get_json()
        
        # Get existing metric
        existing_metric = storage.get_physical_metric(metric_id, user_id)
        if not existing_metric:
            return jsonify({'success': False, 'errors': ['Physical metric not found']}), 404
        
        # Parse date if needed
        date_str = data.get('date', existing_metric.date)
        if isinstance(date_str, str) and '-' in date_str and len(date_str) == 10:
            try:
                date_str = parse_input_date(date_str)
            except Exception:
                pass  # Let the model validator handle it
        
        # Helper functions
        def to_float_or_none(value, default=None):
            if value is None or value == '':
                return default
            try:
                return float(value)
            except (ValueError, TypeError):
                return default
        
        def to_int_or_none(value, default=None):
            if value is None or value == '':
                return default
            try:
                return int(value)
            except (ValueError, TypeError):
                return default
        
        # Auto-convert sprint speed: if m/s is provided, calculate km/h; if only km/h, calculate m/s
        sprint_speed_ms = to_float_or_none(data.get('sprint_speed_ms'), existing_metric.sprint_speed_ms)
        sprint_speed_kmh = to_float_or_none(data.get('sprint_speed_kmh'), existing_metric.sprint_speed_kmh)
        
        if sprint_speed_ms and not sprint_speed_kmh:
            # Calculate km/h from m/s
            sprint_speed_kmh = sprint_speed_ms * 3.6
        elif sprint_speed_kmh and not sprint_speed_ms:
            # Calculate m/s from km/h
            sprint_speed_ms = sprint_speed_kmh / 3.6
        elif sprint_speed_ms and sprint_speed_kmh:
            # Both provided - prefer m/s and recalculate km/h for consistency
            sprint_speed_kmh = sprint_speed_ms * 3.6
        
        # Update metric object
        metric = PhysicalMetrics(
            id=metric_id,
            date=date_str,
            sprint_speed_ms=sprint_speed_ms,
            sprint_speed_kmh=sprint_speed_kmh,
            sprint_10m_sec=to_float_or_none(data.get('sprint_10m_sec'), existing_metric.sprint_10m_sec),
            sprint_20m_sec=to_float_or_none(data.get('sprint_20m_sec'), existing_metric.sprint_20m_sec),
            sprint_30m_sec=to_float_or_none(data.get('sprint_30m_sec'), existing_metric.sprint_30m_sec),
            vertical_jump_cm=to_float_or_none(data.get('vertical_jump_cm'), existing_metric.vertical_jump_cm),
            standing_long_jump_cm=to_float_or_none(data.get('standing_long_jump_cm'), existing_metric.standing_long_jump_cm),
            countermovement_jump_cm=to_float_or_none(data.get('countermovement_jump_cm'), existing_metric.countermovement_jump_cm),
            agility_time_sec=to_float_or_none(data.get('agility_time_sec'), existing_metric.agility_time_sec),
            yo_yo_test_level=to_float_or_none(data.get('yo_yo_test_level'), existing_metric.yo_yo_test_level),
            beep_test_level=to_float_or_none(data.get('beep_test_level'), existing_metric.beep_test_level),
            bench_press_kg=to_float_or_none(data.get('bench_press_kg'), existing_metric.bench_press_kg),
            squat_kg=to_float_or_none(data.get('squat_kg'), existing_metric.squat_kg),
            deadlift_kg=to_float_or_none(data.get('deadlift_kg'), existing_metric.deadlift_kg),
            vo2_max=to_float_or_none(data.get('vo2_max'), existing_metric.vo2_max),
            max_heart_rate=to_int_or_none(data.get('max_heart_rate'), existing_metric.max_heart_rate),
            resting_heart_rate=to_int_or_none(data.get('resting_heart_rate'), existing_metric.resting_heart_rate),
            sit_and_reach_cm=to_float_or_none(data.get('sit_and_reach_cm'), existing_metric.sit_and_reach_cm),
            notes=data.get('notes', existing_metric.notes),
            include_in_report=data.get('include_in_report', getattr(existing_metric, 'include_in_report', True))
        )
        
        # Save updated metric
        storage.save_physical_metric(metric, user_id)
        
        return jsonify({'success': True})
        
    except (ValueError, TypeError, KeyError) as e:
        return jsonify({'success': False, 'errors': [f'Invalid physical metric data: {str(e)}']}), 400
    except Exception as e:
        return jsonify({'success': False, 'errors': [str(e)]}), 400


@bp.route('/api/physical-metrics/<metric_id>', methods=['DELETE'])
@login_required
def delete_physical_metric(metric_id):
    """Delete a physical metric entry"""
    user_id = current_user.id
    success = storage.delete_physical_metric(metric_id, user_id)
    if success:
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'errors': ['Physical metric not found']}), 404


@bp.route('/references')
@login_required
def references_page():
    """References management page - data loaded client-side"""
    # Redirect admin users to admin page
    admin_username = os.environ.get('ADMIN_USERNAME', '').strip()
    if admin_username and current_user.username == admin_username:
        return redirect(url_for('main.admin_users'))
    
    default_settings = {
        'club_name': '',
        'player_name': '',
        'season_year': ''
    }
    return render_template('references.html', 
                         settings=default_settings,
                         references=[])


@bp.route('/admin/users')
@login_required
def admin_users():
    """Admin page to view all registered users"""
    # Simple admin check - you can enhance this with role-based access
    # For now, any logged-in user can view (you may want to restrict this)
    admin_username = os.environ.get('ADMIN_USERNAME', '').strip()
    current_username = current_user.username.strip() if current_user.username else ''
    
    # Debug logging
    current_app.logger.info(f"Admin page access attempt - Admin username: '{admin_username}', Current user: '{current_username}'")
    
    # If ADMIN_USERNAME is set, only that user can access
    if admin_username and current_username != admin_username:
        current_app.logger.warning(f"Access denied - Admin username: '{admin_username}', Current user: '{current_username}'")
        flash('Access denied. Admin access required.', 'error')
        return redirect(url_for('main.dashboard'))
    
    try:
        users = storage.get_all_users()
        # Remove password hashes for security
        users_data = []
        for user in users:
            # Get subscription status for this user
            subscription = storage.get_subscription_by_user_id(user.id)
            subscription_status = subscription.status.value if subscription else 'none'
            subscription_plan = subscription.plan_name if subscription else None
            
            users_data.append({
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'created_at': user.created_at,
                'is_active': user.is_active,
                'subscription_status': subscription_status,
                'subscription_plan': subscription_plan
            })
        
        # Sort by creation date (newest first)
        users_data.sort(key=lambda x: x['created_at'], reverse=True)
        
        return render_template('admin_users.html', users=users_data)
    except Exception as e:
        current_app.logger.error(f"Error loading users for admin: {e}", exc_info=True)
        flash('Error loading users. Please try again.', 'error')
        return redirect(url_for('main.dashboard'))


@bp.route('/api/admin/users')
@login_required
def api_admin_users():
    """API endpoint to get all users as JSON"""
    admin_username = os.environ.get('ADMIN_USERNAME', '').strip()
    
    if admin_username and current_user.username != admin_username:
        return jsonify({'success': False, 'errors': ['Access denied']}), 403
    
    try:
        users = storage.get_all_users()
        users_data = []
        for user in users:
            users_data.append({
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'created_at': user.created_at,
                'is_active': user.is_active
            })
        
        users_data.sort(key=lambda x: x['created_at'], reverse=True)
        
        # Add subscription info to API response
        for user_data in users_data:
            subscription = storage.get_subscription_by_user_id(user_data['id'])
            user_data['subscription_status'] = subscription.status.value if subscription else 'none'
            user_data['subscription_plan'] = subscription.plan_name if subscription else None
        
        return jsonify({'success': True, 'users': users_data, 'total': len(users_data)})
    except Exception as e:
        current_app.logger.error(f"Error loading users for admin API: {e}", exc_info=True)
        return jsonify({'success': False, 'errors': ['Error loading users']}), 500


# ============================================================================
# Legal & Support Pages
# ============================================================================

@bp.route('/privacy')
def privacy():
    """Privacy Policy page"""
    return render_template('legal/privacy.html', 
                         support_email=SUPPORT_EMAIL,
                         current_year=CURRENT_YEAR)

@bp.route('/terms')
def terms():
    """Terms & Conditions page"""
    return render_template('legal/terms.html',
                         support_email=SUPPORT_EMAIL,
                         current_year=CURRENT_YEAR)

@bp.route('/safeguarding')
def safeguarding():
    """Child Safety & Safeguarding page"""
    return render_template('legal/safeguarding.html',
                         support_email=SUPPORT_EMAIL,
                         current_year=CURRENT_YEAR)

@bp.route('/disclaimers')
def disclaimers():
    """Disclaimers page (PHV, Accuracy, etc.)"""
    return render_template('legal/disclaimers.html',
                         support_email=SUPPORT_EMAIL,
                         current_year=CURRENT_YEAR)

@bp.route('/subscription-info')
def subscription_info():
    """Subscription & Billing Information page"""
    return render_template('legal/subscription_info.html',
                         pricing=SUBSCRIPTION_PRICING,
                         support_email=SUPPORT_EMAIL,
                         current_year=CURRENT_YEAR)

@bp.route('/contact')
def contact():
    """Contact & Support page"""
    return render_template('legal/contact.html',
                         support_email=SUPPORT_EMAIL,
                         current_year=CURRENT_YEAR)

@bp.route('/faq')
def faq():
    """Frequently Asked Questions page"""
    return render_template('legal/faq.html',
                         support_email=SUPPORT_EMAIL,
                         current_year=CURRENT_YEAR)

@bp.route('/example-report')
def example_report():
    """Example Report Preview page"""
    return render_template('legal/example_report.html',
                         support_email=SUPPORT_EMAIL,
                         current_year=CURRENT_YEAR)
